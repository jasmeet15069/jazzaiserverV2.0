"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                        JAZZ AI  —  v11.0  Production                        ║
║                                                                              ║
║  Architecture : Modular FastAPI · SQLite/WAL · ChromaDB · APScheduler       ║
║  Auth         : JWT + Refresh tokens · API-keys (scoped) · Rate-limiting     ║
║  Chat         : Session · streaming SSE · branching · edit · regenerate      ║
║  Context      : Sliding-window + rolling summaries + memory injection + RAG  ║
║  RAG          : Chunk → embed → hybrid retrieval → re-rank                  ║
║  Memory       : Auto-extract · tiered priority · confidence scoring          ║
║  Agents       : Think → Act → Observe loop · tool registry · cron jobs      ║
║  Connectors   : Slack · Google Sheets · Excel · Power BI · Generic HTTP      ║
║  LiveKit      : Voice rooms · STT (Whisper) · TTS (PlayAI) · multilingual   ║
║  Thinking     : Auto reasoning mode · chain-of-thought · deep analysis       ║
║  Computer     : Full system access · shell · browser · file ops              ║
║  Admin        : Impersonation · GDPR export · audit log · announcements      ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

# ─── stdlib ───────────────────────────────────────────────────────────────────
import asyncio
import base64
import csv
import hashlib
import io
import json
import logging
import math
import os
import platform
import re
import secrets
import shutil
import subprocess
import sys
import tempfile
import time
import traceback
import uuid
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, List, Optional, Set, Tuple

# ─── third-party ──────────────────────────────────────────────────────────────
import aiosqlite
try:
    import fitz  # PyMuPDF
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.interval import IntervalTrigger
from cryptography.fernet import Fernet

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

from dotenv import load_dotenv
from fastapi import (
    BackgroundTasks, Depends, FastAPI, File, HTTPException,
    Request, UploadFile, status, WebSocket, WebSocketDisconnect,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import (
    FileResponse, HTMLResponse, JSONResponse, StreamingResponse,
)
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from fastapi.staticfiles import StaticFiles
from jose import JWTError, jwt
from openai import OpenAI
from passlib.context import CryptContext
from pydantic import BaseModel, Field, field_validator, model_validator

try:
    import chromadb
    from chromadb.utils import embedding_functions
    HAS_CHROMA = True
except ImportError:
    HAS_CHROMA = False

load_dotenv()

# ══════════════════════════════════════════════════════════════════════════════
# §1  LOGGING
# ══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    handlers=[
        logging.FileHandler("jazz.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger("jazz")


# ══════════════════════════════════════════════════════════════════════════════
# §2  CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

def _env(key: str, default: str = "") -> str:
    val = os.getenv(key, default)
    if not val:
        logger.warning("[CONFIG] %s not set", key)
    return val


# ── JWT ───────────────────────────────────────────────────────────────────────
JWT_SECRET: str = os.getenv("JWT_SECRET_KEY") or secrets.token_urlsafe(64)
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES  = 60 * 24 * 7
REFRESH_TOKEN_EXPIRE_MINUTES = 60 * 24 * 30

# ── Fernet ────────────────────────────────────────────────────────────────────
_FERNET_KEY_RAW = os.getenv("FERNET_KEY", "") or Fernet.generate_key().decode()
_fernet = Fernet(
    _FERNET_KEY_RAW.encode() if isinstance(_FERNET_KEY_RAW, str) else _FERNET_KEY_RAW
)

def _encrypt(data: dict) -> str:
    return _fernet.encrypt(json.dumps(data).encode()).decode()

def _decrypt(token: str) -> dict:
    return json.loads(_fernet.decrypt(token.encode()).decode())

# ── Service keys ──────────────────────────────────────────────────────────────
ADMIN_EMAIL    = os.getenv("ADMIN_EMAIL", "jasmeet.15069@gmail.com")
ADMIN_PASSWORD = _env("ADMIN_PASSWORD","Acx@POWER@12345jassi789")
GROQ_API_KEY   = _env("GROQ_API_KEY")
HF_TOKEN       = _env("HF_TOKEN")
SLACK_BOT_TOKEN = os.getenv("SLACK_BOT_TOKEN", "")

# ── LiveKit ───────────────────────────────────────────────────────────────────
LIVEKIT_URL        = os.getenv("LIVEKIT_URL", "wss://jazzbot-7d7ucr5z.livekit.cloud")
LIVEKIT_API_KEY    = os.getenv("LIVEKIT_API_KEY", "APISUmh3EtLiLyb")
LIVEKIT_API_SECRET = os.getenv("LIVEKIT_API_SECRET", "4hm7cffNQBYnkRiOYOrwfxgMsi5EpHIn8kJeT1mttYsA")
TTS_VOICE          = os.getenv("TTS_VOICE", "Fritz-PlayAI")

# ── Storage ───────────────────────────────────────────────────────────────────
DB_PATH       = str(Path(os.getenv("DB_PATH", "./jazz_v11.db")).resolve())
WORKSPACE_DIR = Path("workspace"); WORKSPACE_DIR.mkdir(exist_ok=True)
SITES_DIR     = WORKSPACE_DIR / "sites";   SITES_DIR.mkdir(exist_ok=True)
UPLOADS_DIR   = WORKSPACE_DIR / "uploads"; UPLOADS_DIR.mkdir(exist_ok=True)

# ── RAG / context ─────────────────────────────────────────────────────────────
CHUNK_SIZE        = 500
CHUNK_OVERLAP     = 50
TOP_K_RETRIEVAL   = 6
RERANK_TOP_K      = 3
MAX_CONTEXT_TOKENS = 6000
SUMMARY_TRIGGER   = 20

# ── Thinking mode thresholds ──────────────────────────────────────────────────
THINKING_KEYWORDS = [
    "why", "how does", "explain", "analyze", "reason", "think", "understand",
    "compare", "difference", "pros and cons", "debate", "argue", "philosophy",
    "ethics", "theory", "hypothesis", "research", "complex", "deep dive",
    "step by step", "break down", "implications", "consequences", "tradeoffs",
    "solve", "algorithm", "proof", "derive", "calculate", "optimize",
    "should i", "what if", "is it true", "help me understand", "critically",
    "क्यों", "कैसे", "समझाओ", "विश्लेषण", "सोचो",
]
THINKING_MIN_LENGTH = 50

# ── Rate / limits ─────────────────────────────────────────────────────────────
MAX_FILE_SIZE   = 50  * 1024 * 1024
MAX_IMAGE_SIZE  = 20  * 1024 * 1024
CODE_EXEC_TIMEOUT = 30
CODE_EXEC_MAX_OUT = 100_000
MAX_MEMORIES    = 200

ALLOWED_ORIGINS: List[str] = [
    o.strip() for o in os.getenv("ALLOWED_ORIGINS", "*").split(",") if o.strip()
]
ALLOWED_EXTENSIONS = {
    ".html", ".css", ".js", ".py", ".txt", ".md", ".json", ".xml",
    ".yaml", ".yml", ".csv", ".sql", ".sh", ".ts", ".tsx", ".jsx",
    ".conf", ".ini", ".log", ".svg", ".bat", ".ps1", ".rb", ".go",
    ".rs", ".cpp", ".c", ".h", ".java", ".kt", ".swift", ".php",
}
ALLOWED_DOC_MIMES = {
    "application/pdf", "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/plain", "text/csv", "text/markdown",
}

SUBSCRIPTION_LIMITS: Dict[str, Dict[str, Any]] = {
    "free": {
        "messages_per_day": 100, "documents": 10,  "max_file_mb": 10,
        "agent_jobs": 3,         "websites": 5,    "code_runs_per_day": 50,
        "api_keys": 3,           "chat_sessions": -1, "memories": 100,
        "connectors": 3,
    },
    "pro": {
        "messages_per_day": -1,  "documents": 100,  "max_file_mb": 50,
        "agent_jobs": 20,        "websites": 100,   "code_runs_per_day": -1,
        "api_keys": 20,          "chat_sessions": -1, "memories": -1,
        "connectors": 20,
    },
    "enterprise": {
        "messages_per_day": -1,  "documents": -1,  "max_file_mb": 200,
        "agent_jobs": -1,        "websites": -1,   "code_runs_per_day": -1,
        "api_keys": -1,          "chat_sessions": -1, "memories": -1,
        "connectors": -1,
    },
}

_SERVER_START = datetime.now(timezone.utc)
_executor     = ThreadPoolExecutor(max_workers=16)


# ══════════════════════════════════════════════════════════════════════════════
# §3  DATABASE
# ══════════════════════════════════════════════════════════════════════════════

_db: Optional[aiosqlite.Connection] = None


async def _init_db_connection() -> None:
    global _db
    _db = await aiosqlite.connect(DB_PATH, check_same_thread=False)
    _db.row_factory = aiosqlite.Row
    for pragma in [
        "PRAGMA journal_mode=WAL",
        "PRAGMA synchronous=NORMAL",
        "PRAGMA foreign_keys=ON",
        "PRAGMA cache_size=-32000",
        "PRAGMA temp_store=MEMORY",
        "PRAGMA mmap_size=268435456",
    ]:
        await _db.execute(pragma)
    await _db.commit()


def _conn() -> aiosqlite.Connection:
    if _db is None:
        raise RuntimeError("DB not initialised")
    return _db


async def db_fetchone(sql: str, params: tuple = ()) -> Optional[Dict]:
    async with _conn().execute(sql, params) as cur:
        row = await cur.fetchone()
        return dict(row) if row else None

async def db_fetchall(sql: str, params: tuple = ()) -> List[Dict]:
    async with _conn().execute(sql, params) as cur:
        rows = await cur.fetchall()
        return [dict(r) for r in rows]

async def db_count(sql: str, params: tuple = ()) -> int:
    row = await db_fetchone(sql, params)
    return list(row.values())[0] if row else 0

async def db_execute(sql: str, params: tuple = ()) -> int:
    async with _conn().execute(sql, params) as cur:
        await _conn().commit()
        return cur.lastrowid or 0

async def db_executemany(sql: str, param_list: List[tuple]) -> None:
    await _conn().executemany(sql, param_list)
    await _conn().commit()

async def db_script(ddl: str) -> None:
    await _conn().executescript(ddl)

@asynccontextmanager
async def db_transaction():
    await _conn().execute("BEGIN")
    try:
        yield
        await _conn().commit()
    except Exception:
        await _conn().rollback()
        raise


# ── Schema ────────────────────────────────────────────────────────────────────
_SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id TEXT PRIMARY KEY, email TEXT NOT NULL UNIQUE,
    password_hash TEXT NOT NULL, full_name TEXT DEFAULT '',
    role TEXT DEFAULT 'client' CHECK(role IN('admin','client')),
    subscription TEXT DEFAULT 'free' CHECK(subscription IN('free','pro','enterprise')),
    subscription_expires_at TEXT, memory_enabled INTEGER DEFAULT 1,
    timezone TEXT DEFAULT 'UTC', is_active INTEGER DEFAULT 1, is_verified INTEGER DEFAULT 0,
    last_login_at TEXT, created_at TEXT NOT NULL, updated_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS refresh_tokens (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL, token_hash TEXT NOT NULL UNIQUE,
    expires_at TEXT NOT NULL, revoked INTEGER DEFAULT 0, created_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS chat_sessions (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL, title TEXT DEFAULT 'New Chat',
    model_id TEXT DEFAULT 'censored', is_pinned INTEGER DEFAULT 0, is_archived INTEGER DEFAULT 0,
    turn_count INTEGER DEFAULT 0, last_message_at TEXT,
    parent_session_id TEXT, branch_from_msg_id TEXT,
    created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS chat_history (
    id TEXT PRIMARY KEY, session_id TEXT NOT NULL, user_id TEXT NOT NULL,
    role TEXT NOT NULL CHECK(role IN('user','assistant','system','tool')),
    content TEXT NOT NULL, model_used TEXT, agent_used TEXT,
    thinking_content TEXT DEFAULT '',
    tool_calls_json TEXT DEFAULT '[]', rag_chunks_json TEXT DEFAULT '[]',
    tokens_input INTEGER DEFAULT 0, tokens_output INTEGER DEFAULT 0,
    latency_ms INTEGER DEFAULT 0, is_hidden INTEGER DEFAULT 0,
    parent_msg_id TEXT, edit_count INTEGER DEFAULT 0,
    mode TEXT DEFAULT 'normal',
    created_at TEXT NOT NULL,
    FOREIGN KEY(session_id) REFERENCES chat_sessions(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS message_edits (
    id TEXT PRIMARY KEY, message_id TEXT NOT NULL,
    old_content TEXT NOT NULL, new_content TEXT NOT NULL, edited_at TEXT NOT NULL,
    FOREIGN KEY(message_id) REFERENCES chat_history(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS summaries (
    id TEXT PRIMARY KEY, session_id TEXT NOT NULL, user_id TEXT NOT NULL,
    summary_text TEXT NOT NULL, covers_from_at TEXT NOT NULL, covers_to_at TEXT NOT NULL,
    turn_count INTEGER NOT NULL, tokens_saved INTEGER DEFAULT 0,
    model_used TEXT NOT NULL, created_at TEXT NOT NULL,
    FOREIGN KEY(session_id) REFERENCES chat_sessions(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS memories (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
    key TEXT NOT NULL, value TEXT NOT NULL,
    source TEXT DEFAULT 'auto' CHECK(source IN('auto','manual','agent')),
    confidence REAL DEFAULT 1.0, last_reinforced TEXT NOT NULL,
    reinforcement_count INTEGER DEFAULT 1, is_active INTEGER DEFAULT 1,
    created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
    UNIQUE(user_id, key),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS documents (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
    filename TEXT NOT NULL, original_name TEXT NOT NULL, file_type TEXT DEFAULT '',
    file_size_bytes INTEGER DEFAULT 0, page_count INTEGER, chunk_count INTEGER DEFAULT 0,
    char_count INTEGER DEFAULT 0, is_indexed INTEGER DEFAULT 0, index_error TEXT,
    collection_name TEXT DEFAULT 'documents', tags_json TEXT DEFAULT '[]',
    description TEXT DEFAULT '', uploaded_at TEXT NOT NULL, indexed_at TEXT,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS embeddings_meta (
    id TEXT PRIMARY KEY, document_id TEXT NOT NULL, user_id TEXT NOT NULL,
    chunk_index INTEGER NOT NULL, chunk_text TEXT NOT NULL,
    chroma_id TEXT NOT NULL UNIQUE, embed_model TEXT NOT NULL,
    token_count INTEGER DEFAULT 0, created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS api_keys (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL, name TEXT NOT NULL,
    key_hash TEXT NOT NULL UNIQUE, key_prefix TEXT NOT NULL,
    scopes_json TEXT DEFAULT '["chat","rag"]',
    rate_limit_rpm INTEGER DEFAULT 60, is_active INTEGER DEFAULT 1,
    last_used_at TEXT, usage_count INTEGER DEFAULT 0, expires_at TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS agent_jobs (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL, name TEXT NOT NULL,
    description TEXT DEFAULT '', job_type TEXT NOT NULL,
    trigger_json TEXT DEFAULT '{}', prompt_template TEXT NOT NULL,
    tools_json TEXT DEFAULT '[]', model_id TEXT DEFAULT 'censored',
    status TEXT DEFAULT 'idle', enabled INTEGER DEFAULT 1,
    max_retries INTEGER DEFAULT 2, retry_count INTEGER DEFAULT 0,
    timeout_seconds INTEGER DEFAULT 120,
    total_runs INTEGER DEFAULT 0, success_runs INTEGER DEFAULT 0, failed_runs INTEGER DEFAULT 0,
    last_run_at TEXT, last_run_status TEXT, next_run_at TEXT,
    metadata_json TEXT DEFAULT '{}', created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS agent_job_logs (
    id TEXT PRIMARY KEY, job_id TEXT NOT NULL, user_id TEXT NOT NULL,
    run_number INTEGER NOT NULL, status TEXT NOT NULL, trigger_type TEXT DEFAULT 'cron',
    prompt_rendered TEXT NOT NULL, result_text TEXT,
    tool_calls_json TEXT DEFAULT '[]', tokens_used INTEGER DEFAULT 0,
    latency_ms INTEGER DEFAULT 0, error_message TEXT,
    started_at TEXT NOT NULL, finished_at TEXT,
    FOREIGN KEY(job_id) REFERENCES agent_jobs(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS connectors (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL, name TEXT NOT NULL,
    connector_type TEXT NOT NULL, encrypted_creds TEXT NOT NULL,
    is_active INTEGER DEFAULT 1, last_tested_at TEXT,
    last_test_ok INTEGER, test_error TEXT,
    metadata_json TEXT DEFAULT '{}', created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
    UNIQUE(user_id, name),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS notifications (
    id TEXT PRIMARY KEY, user_id TEXT,
    title TEXT NOT NULL, message TEXT NOT NULL,
    type TEXT DEFAULT 'info', link TEXT,
    is_read INTEGER DEFAULT 0, is_broadcast INTEGER DEFAULT 0,
    actor_id TEXT, expires_at TEXT, created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS coupons (
    id TEXT PRIMARY KEY, code TEXT NOT NULL UNIQUE COLLATE NOCASE,
    grants_tier TEXT NOT NULL, duration_days INTEGER DEFAULT 30,
    max_uses INTEGER DEFAULT 1, current_uses INTEGER DEFAULT 0,
    is_active INTEGER DEFAULT 1, created_by TEXT NOT NULL,
    expires_at TEXT, created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS coupon_redemptions (
    id TEXT PRIMARY KEY, coupon_id TEXT NOT NULL, user_id TEXT NOT NULL,
    redeemed_at TEXT NOT NULL, UNIQUE(coupon_id, user_id)
);
CREATE TABLE IF NOT EXISTS websites (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
    title TEXT NOT NULL, description TEXT DEFAULT '',
    filename TEXT NOT NULL UNIQUE, style TEXT DEFAULT 'modern',
    prompt_used TEXT, html_size_bytes INTEGER DEFAULT 0,
    is_public INTEGER DEFAULT 0, view_count INTEGER DEFAULT 0,
    created_at TEXT NOT NULL, updated_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS audit_log (
    id TEXT PRIMARY KEY, actor_id TEXT NOT NULL, target_id TEXT,
    action TEXT NOT NULL, detail_json TEXT DEFAULT '{}',
    ip_address TEXT, created_at TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS rate_limits (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
    resource TEXT NOT NULL, window_key TEXT NOT NULL,
    count INTEGER DEFAULT 0, updated_at TEXT NOT NULL,
    UNIQUE(user_id, resource, window_key)
);
CREATE TABLE IF NOT EXISTS message_feedback (
    id TEXT PRIMARY KEY, message_id TEXT NOT NULL, user_id TEXT NOT NULL,
    rating INTEGER DEFAULT 0, feedback TEXT DEFAULT '',
    created_at TEXT NOT NULL, UNIQUE(message_id, user_id)
);
CREATE TABLE IF NOT EXISTS voice_sessions (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
    room_name TEXT NOT NULL, livekit_token TEXT,
    language TEXT DEFAULT 'auto', status TEXT DEFAULT 'active',
    started_at TEXT NOT NULL, ended_at TEXT,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS computer_tasks (
    id TEXT PRIMARY KEY, user_id TEXT NOT NULL,
    task_type TEXT NOT NULL, command TEXT NOT NULL,
    result TEXT, status TEXT DEFAULT 'pending',
    stdout TEXT, stderr TEXT, exit_code INTEGER,
    created_at TEXT NOT NULL, completed_at TEXT,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE INDEX IF NOT EXISTS idx_users_email          ON users(email);
CREATE INDEX IF NOT EXISTS idx_sessions_user        ON chat_sessions(user_id, last_message_at DESC);
CREATE INDEX IF NOT EXISTS idx_history_session      ON chat_history(session_id, created_at);
CREATE INDEX IF NOT EXISTS idx_memories_user        ON memories(user_id, is_active, last_reinforced DESC);
CREATE INDEX IF NOT EXISTS idx_documents_user       ON documents(user_id, uploaded_at DESC);
CREATE INDEX IF NOT EXISTS idx_api_keys_hash        ON api_keys(key_hash);
CREATE INDEX IF NOT EXISTS idx_agent_jobs_user      ON agent_jobs(user_id, enabled);
CREATE INDEX IF NOT EXISTS idx_job_logs_job         ON agent_job_logs(job_id, started_at DESC);
CREATE INDEX IF NOT EXISTS idx_connectors_user      ON connectors(user_id, is_active);
CREATE INDEX IF NOT EXISTS idx_notifications_user   ON notifications(user_id, is_read, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_rate_limits          ON rate_limits(user_id, resource, window_key);
CREATE INDEX IF NOT EXISTS idx_refresh_tokens       ON refresh_tokens(token_hash, revoked);
"""

async def _bootstrap_schema() -> None:
    await db_script(_SCHEMA)
    logger.info("[DB] Schema applied ✅")

def _truncate_password(pwd: str) -> str:
    return pwd.encode('utf-8')[:72].decode('utf-8', errors='ignore')

async def _bootstrap_admin() -> None:
    existing = await db_fetchone("SELECT id FROM users WHERE email=?", (ADMIN_EMAIL,))
    if not existing and ADMIN_PASSWORD:
        uid = str(uuid.uuid4()); now = _utcnow()
        ph = _pwd.hash(_truncate_password(ADMIN_PASSWORD or "Admin@Jazz123!"))
        await db_execute(
            "INSERT INTO users (id,email,password_hash,full_name,role,subscription,created_at,updated_at)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (uid, ADMIN_EMAIL, ph, "Administrator", "admin", "enterprise", now, now),
        )
        logger.info("[DB] Admin bootstrapped: %s", ADMIN_EMAIL)

def _utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()

def _new_id() -> str:
    return str(uuid.uuid4())


# ══════════════════════════════════════════════════════════════════════════════
# §4  AUTH
# ══════════════════════════════════════════════════════════════════════════════

_pwd      = CryptContext(schemes=["bcrypt"], deprecated="auto")
_security = HTTPBearer(auto_error=False)

def _hash_token(raw: str) -> str:
    return hashlib.sha256(raw.encode()).hexdigest()

def _create_jwt(data: dict, expire_minutes: int) -> str:
    exp = datetime.now(timezone.utc) + timedelta(minutes=expire_minutes)
    return jwt.encode({**data, "exp": exp}, JWT_SECRET, algorithm=JWT_ALGORITHM)

def _decode_jwt(token: str) -> Optional[Dict]:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except JWTError:
        return None

async def _create_refresh_token(user_id: str) -> str:
    raw = f"jzr_{secrets.token_urlsafe(48)}"
    h   = _hash_token(raw)
    exp = (datetime.now(timezone.utc) + timedelta(minutes=REFRESH_TOKEN_EXPIRE_MINUTES)).isoformat()
    await db_execute(
        "INSERT INTO refresh_tokens (id,user_id,token_hash,expires_at,created_at) VALUES (?,?,?,?,?)",
        (_new_id(), user_id, h, exp, _utcnow()),
    )
    return raw

async def _validate_refresh_token(raw: str) -> Optional[str]:
    h   = _hash_token(raw)
    row = await db_fetchone(
        "SELECT user_id, expires_at FROM refresh_tokens WHERE token_hash=? AND revoked=0", (h,)
    )
    if not row:
        return None
    if datetime.fromisoformat(row["expires_at"]) < datetime.now(timezone.utc):
        return None
    return row["user_id"]

async def _revoke_refresh_token(raw: str) -> None:
    h = _hash_token(raw)
    await db_execute("UPDATE refresh_tokens SET revoked=1 WHERE token_hash=?", (h,))

async def _auth_via_api_key(raw_key: str) -> Optional[Dict]:
    if not raw_key.startswith("jz_"):
        return None
    h   = _hash_token(raw_key)
    row = await db_fetchone(
        "SELECT ak.*, u.email, u.role, u.subscription FROM api_keys ak"
        " JOIN users u ON ak.user_id=u.id"
        " WHERE ak.key_hash=? AND ak.is_active=1", (h,),
    )
    if not row:
        return None
    if row.get("expires_at") and datetime.fromisoformat(row["expires_at"]) < datetime.now(timezone.utc):
        return None
    asyncio.create_task(db_execute(
        "UPDATE api_keys SET last_used_at=?, usage_count=usage_count+1 WHERE id=?",
        (_utcnow(), row["id"]),
    ))
    return {
        "sub": row["user_id"], "email": row["email"],
        "role": row["role"], "subscription": row["subscription"],
        "scopes": json.loads(row.get("scopes_json", '["chat"]')),
        "auth_method": "api_key",
    }

async def _resolve_auth(request: Request, credentials: Optional[HTTPAuthorizationCredentials]) -> Optional[Dict]:
    if credentials and credentials.credentials:
        tok = credentials.credentials
        if tok.startswith("jz_"):
            return await _auth_via_api_key(tok)
        payload = _decode_jwt(tok)
        if payload:
            return payload
    ak = request.headers.get("X-API-Key", "")
    if ak.startswith("jz_"):
        return await _auth_via_api_key(ak)
    qt = request.query_params.get("token", "")
    if qt:
        if qt.startswith("jz_"):
            return await _auth_via_api_key(qt)
        payload = _decode_jwt(qt)
        if payload:
            return payload
    return None

async def get_current_user(request: Request, credentials: Optional[HTTPAuthorizationCredentials] = Depends(_security)) -> Dict:
    payload = await _resolve_auth(request, credentials)
    if not payload:
        raise HTTPException(401, "Authorization required", headers={"WWW-Authenticate": "Bearer"})
    return payload

async def require_admin(request: Request, credentials: Optional[HTTPAuthorizationCredentials] = Depends(_security)) -> Dict:
    payload = await _resolve_auth(request, credentials)
    if not payload:
        raise HTTPException(401, "Authorization required")
    if payload.get("role") != "admin":
        raise HTTPException(403, "Admin access required")
    return payload

def _require_scope(scope: str):
    async def _guard(user: Dict = Depends(get_current_user)) -> Dict:
        scopes = user.get("scopes")
        if scopes is not None and scope not in scopes:
            raise HTTPException(403, f"API key missing scope: {scope}")
        return user
    return _guard

class PersistentRateLimiter:
    def _window_key(self) -> str:
        return datetime.now(timezone.utc).strftime("%Y-%m-%d")

    async def check(self, user_id: str, subscription: str, resource: str) -> bool:
        limit = SUBSCRIPTION_LIMITS.get(subscription, SUBSCRIPTION_LIMITS["free"]).get(resource, 50)
        if limit == -1:
            return True
        wk  = self._window_key()
        row = await db_fetchone(
            "SELECT count FROM rate_limits WHERE user_id=? AND resource=? AND window_key=?",
            (user_id, resource, wk),
        )
        current = row["count"] if row else 0
        if current >= limit:
            return False
        if row:
            await db_execute(
                "UPDATE rate_limits SET count=count+1, updated_at=? WHERE user_id=? AND resource=? AND window_key=?",
                (_utcnow(), user_id, resource, wk),
            )
        else:
            await db_execute(
                "INSERT OR IGNORE INTO rate_limits (id,user_id,resource,window_key,count,updated_at)"
                " VALUES (?,?,?,?,1,?)",
                (_new_id(), user_id, resource, wk, _utcnow()),
            )
        return True

    async def usage(self, user_id: str, resource: str) -> int:
        wk  = self._window_key()
        row = await db_fetchone(
            "SELECT count FROM rate_limits WHERE user_id=? AND resource=? AND window_key=?",
            (user_id, resource, wk),
        )
        return row["count"] if row else 0

rate_limiter = PersistentRateLimiter()

async def _check_rate(user: Dict, resource: str) -> None:
    ok = await rate_limiter.check(user["sub"], user.get("subscription", "free"), resource)
    if not ok:
        raise HTTPException(429, f"Daily {resource} limit reached. Upgrade your plan.")


# ══════════════════════════════════════════════════════════════════════════════
# §5  CHROMADB
# ══════════════════════════════════════════════════════════════════════════════

_chroma_client: Any = None
_docs_collection: Any = None
_schema_collection: Any = None
_embed_fn: Any = None

def _init_chroma() -> None:
    global _chroma_client, _docs_collection, _schema_collection, _embed_fn
    if not HAS_CHROMA:
        logger.warning("[CHROMA] chromadb not installed — RAG disabled")
        return
    try:
        _embed_fn = embedding_functions.SentenceTransformerEmbeddingFunction(
            model_name="all-MiniLM-L6-v2"
        )
        _chroma_client = chromadb.PersistentClient(path="./chroma_db")
        _docs_collection = _chroma_client.get_or_create_collection(
            name="documents", embedding_function=_embed_fn,
            metadata={"hnsw:space": "cosine"},
        )
        _schema_collection = _chroma_client.get_or_create_collection(
            name="db_schema", embedding_function=_embed_fn,
        )
        logger.info("[CHROMA] Initialized ✅  docs=%d", _docs_collection.count())
    except Exception as exc:
        logger.error("[CHROMA] Init failed: %s ❌", exc)


# ══════════════════════════════════════════════════════════════════════════════
# §6  LLM ROUTER
# ══════════════════════════════════════════════════════════════════════════════

_MODEL_REGISTRY: Dict[str, Dict[str, str]] = {
    "censored": {
        "base_url": "https://api.groq.com/openai/v1",
        "api_key":  GROQ_API_KEY,
        "model":    "llama-3.3-70b-versatile",
        "label":    "Groq Llama 3.3 70B",
    },
    "uncensored": {
        "base_url": "https://router.huggingface.co/v1",
        "api_key":  HF_TOKEN,
        "model":    "dphn/Dolphin-Mistral-24B-Venice-Edition:featherless-ai",
        "label":    "Dolphin Mistral 24B",
    },
    "fast": {
        "base_url": "https://api.groq.com/openai/v1",
        "api_key":  GROQ_API_KEY,
        "model":    "llama-3.1-8b-instant",
        "label":    "Groq Llama 3.1 8B (fast)",
    },
    "thinking": {
        "base_url": "https://api.groq.com/openai/v1",
        "api_key":  GROQ_API_KEY,
        "model":    "llama-3.3-70b-versatile",
        "label":    "Thinking / Reasoning Mode",
    },
}
_VISION_MODEL = {
    "base_url": "https://router.huggingface.co/v1",
    "api_key":  HF_TOKEN,
    "model":    "Qwen/Qwen2.5-VL-7B-Instruct:fastest",
}
_FALLBACK_CHAIN = ["censored", "fast"]


def _get_llm_client(model_id: str) -> Tuple[OpenAI, str]:
    cfg = _MODEL_REGISTRY.get(model_id) or _MODEL_REGISTRY["censored"]
    return OpenAI(base_url=cfg["base_url"], api_key=cfg["api_key"]), cfg["model"]

def _get_vision_client() -> Tuple[OpenAI, str]:
    cfg = _VISION_MODEL
    return OpenAI(base_url=cfg["base_url"], api_key=cfg["api_key"]), cfg["model"]

async def _llm_text(
    messages: List[Dict], model_id: str = "censored",
    max_tokens: int = 1024, temperature: float = 0.7,
) -> str:
    chain = [model_id] + [m for m in _FALLBACK_CHAIN if m != model_id]
    last_exc: Exception = RuntimeError("No models")
    for mid in chain:
        try:
            client, model = _get_llm_client(mid)
            resp = await asyncio.get_event_loop().run_in_executor(
                _executor,
                lambda c=client, m=model: c.chat.completions.create(
                    model=m, messages=messages, max_tokens=max_tokens, temperature=temperature
                ),
            )
            return resp.choices[0].message.content or ""
        except Exception as exc:
            logger.warning("[LLM] %s failed: %s", mid, exc)
            last_exc = exc
    raise last_exc


# ══════════════════════════════════════════════════════════════════════════════
# §7  THINKING / REASONING MODE
# ══════════════════════════════════════════════════════════════════════════════

def _should_think(message: str) -> bool:
    """Auto-detect if message needs thinking/reasoning mode."""
    msg_lower = message.lower().strip()
    if len(message) < THINKING_MIN_LENGTH:
        return False
    if any(kw in msg_lower for kw in THINKING_KEYWORDS):
        return True
    # Complex questions: many words + question mark
    word_count = len(message.split())
    if word_count > 30 and "?" in message:
        return True
    # Code complexity signals
    code_signals = ["debug", "fix", "implement", "build", "create", "architect", "design system"]
    if any(s in msg_lower for s in code_signals) and word_count > 20:
        return True
    return False


_THINKING_SYSTEM = """You are JAZZ — an advanced AI with deep reasoning capabilities.
When given complex questions, you MUST first think through the problem carefully before answering.

Format your response as:

<thinking>
[Your step-by-step reasoning process here. Be thorough. Explore different angles. 
Consider edge cases. Work through the logic systematically.]
</thinking>

<answer>
[Your final, well-reasoned response here. Be clear, precise, and comprehensive.]
</answer>

Use markdown in your answer for clarity."""

_NORMAL_SYSTEM = """You are JAZZ — a sharp, capable, production-grade AI assistant.
Be direct, precise, and genuinely helpful. Use markdown where it aids clarity.
If you used retrieved document context, cite the source inline.
You have access to the user's computer/system when they ask for system tasks."""


async def _llm_thinking(
    messages: List[Dict], model_id: str = "censored",
    max_tokens: int = 4096, temperature: float = 0.6,
) -> Tuple[str, str]:
    """Returns (thinking_content, final_answer)."""
    # Inject thinking system prompt
    msgs = [{"role": "system", "content": _THINKING_SYSTEM}] + [m for m in messages if m["role"] != "system"]
    chain = [model_id] + [m for m in _FALLBACK_CHAIN if m != model_id]
    last_exc: Exception = RuntimeError("No models")
    for mid in chain:
        try:
            client, model = _get_llm_client(mid)
            resp = await asyncio.get_event_loop().run_in_executor(
                _executor,
                lambda c=client, m=model: c.chat.completions.create(
                    model=m, messages=msgs, max_tokens=max_tokens, temperature=temperature
                ),
            )
            content = resp.choices[0].message.content or ""
            # Parse thinking vs answer
            thinking_match = re.search(r'<thinking>(.*?)</thinking>', content, re.DOTALL)
            answer_match   = re.search(r'<answer>(.*?)</answer>', content, re.DOTALL)
            thinking = thinking_match.group(1).strip() if thinking_match else ""
            answer   = answer_match.group(1).strip() if answer_match else content
            return thinking, answer
        except Exception as exc:
            logger.warning("[THINKING] %s failed: %s", mid, exc)
            last_exc = exc
    raise last_exc


# ══════════════════════════════════════════════════════════════════════════════
# §8  TOKEN COUNTING
# ══════════════════════════════════════════════════════════════════════════════

def _count_tokens(text: str) -> int:
    return math.ceil(len(text) / 4)

def _msgs_tokens(messages: List[Dict]) -> int:
    return sum(_count_tokens(m.get("content", "") or "") for m in messages)


# ══════════════════════════════════════════════════════════════════════════════
# §9  MEMORY SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

async def _get_active_memories(user_id: str, limit: int = 30) -> List[Dict]:
    return await db_fetchall(
        """SELECT key, value, source, confidence, reinforcement_count, last_reinforced
           FROM memories WHERE user_id=? AND is_active=1
           ORDER BY
             CASE source WHEN 'manual' THEN 1 WHEN 'agent' THEN 2 ELSE 3 END,
             (confidence * reinforcement_count) DESC, last_reinforced DESC
           LIMIT ?""",
        (user_id, limit),
    )

def _format_memories(memories: List[Dict]) -> str:
    if not memories:
        return ""
    lines = ["[User context from memory:]"]
    for m in memories:
        src = "✎" if m["source"] == "manual" else "◎"
        lines.append(f"  {src} {m['key']}: {m['value']}")
    return "\n".join(lines)

async def _upsert_memory(user_id: str, key: str, value: str, source: str = "auto", confidence: float = 0.9) -> None:
    now  = _utcnow()
    row  = await db_fetchone(
        "SELECT id, reinforcement_count FROM memories WHERE user_id=? AND key=?", (user_id, key),
    )
    if row:
        new_conf = min(1.0, confidence + 0.05 * row["reinforcement_count"])
        await db_execute(
            "UPDATE memories SET value=?, confidence=?, reinforcement_count=reinforcement_count+1,"
            " last_reinforced=?, updated_at=? WHERE id=?",
            (value, new_conf, now, now, row["id"]),
        )
    else:
        current_count = await db_count(
            "SELECT COUNT(*) FROM memories WHERE user_id=? AND is_active=1", (user_id,)
        )
        if current_count >= MAX_MEMORIES:
            oldest = await db_fetchone(
                "SELECT id FROM memories WHERE user_id=? AND source='auto' AND is_active=1"
                " ORDER BY confidence ASC, last_reinforced ASC LIMIT 1", (user_id,),
            )
            if oldest:
                await db_execute("UPDATE memories SET is_active=0 WHERE id=?", (oldest["id"],))
        await db_execute(
            "INSERT INTO memories (id,user_id,key,value,source,confidence,"
            " last_reinforced,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
            (_new_id(), user_id, key, value, source, confidence, now, now, now),
        )

async def _extract_and_save_memories(user_id: str, user_msg: str, ai_resp: str) -> None:
    try:
        mem_enabled = await db_fetchone("SELECT memory_enabled FROM users WHERE id=?", (user_id,))
        if not mem_enabled or not mem_enabled.get("memory_enabled"):
            return
        prompt = (
            "You are a memory extraction system.\n"
            "Extract 0-4 memorable facts about the USER from this conversation.\n"
            f"User said: {user_msg[:600]}\nAI replied: {ai_resp[:400]}\n\n"
            'Return ONLY valid JSON array: [{"key":"snake_case_key","value":"fact","confidence":0.95}]\n'
            "If nothing memorable, return []."
        )
        result = await _llm_text(
            [{"role": "user", "content": prompt}],
            model_id="fast", max_tokens=300, temperature=0.1,
        )
        raw   = re.sub(r"```json|```", "", result.strip()).strip()
        items = json.loads(raw)
        if not isinstance(items, list):
            return
        for item in items[:4]:
            if not isinstance(item, dict):
                continue
            key   = str(item.get("key", "")).strip()[:40]
            value = str(item.get("value", "")).strip()[:300]
            conf  = float(item.get("confidence", 0.8))
            if key and value:
                await _upsert_memory(user_id, key, value, "auto", conf)
    except Exception as exc:
        logger.debug("[MEMORY] Extraction: %s", exc)


# ══════════════════════════════════════════════════════════════════════════════
# §10  RAG SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

def _extract_pdf(path: str) -> Tuple[str, int]:
    if not HAS_FITZ:
        return "", 0
    try:
        with fitz.open(path) as pdf:
            text = "\n".join(p.get_text() for p in pdf)
            return text, len(pdf)
    except Exception as exc:
        logger.error("[RAG] PDF extract: %s", exc)
        return "", 0

def _extract_docx(path: str) -> str:
    if not HAS_DOCX:
        return ""
    try:
        doc = DocxDocument(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception as exc:
        logger.error("[RAG] DOCX extract: %s", exc)
        return ""

def _extract_xlsx(path: str) -> str:
    if not HAS_OPENPYXL:
        return ""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        parts: List[str] = []
        for sheet in wb.worksheets:
            parts.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_str = " | ".join(str(c) for c in row if c is not None)
                if row_str.strip():
                    parts.append(row_str)
        return "\n".join(parts)
    except Exception as exc:
        logger.error("[RAG] XLSX extract: %s", exc)
        return ""

def _extract_csv(path: str) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return "\n".join(" | ".join(row) for row in csv.reader(f))
    except Exception as exc:
        logger.error("[RAG] CSV extract: %s", exc)
        return ""

def _parse_document(path: str, mime: str) -> Tuple[str, int]:
    if "pdf" in mime:
        return _extract_pdf(path)
    if "word" in mime or path.endswith(".docx"):
        return _extract_docx(path), 0
    if "excel" in mime or path.endswith((".xlsx", ".xls")):
        return _extract_xlsx(path), 0
    if "csv" in mime or path.endswith(".csv"):
        return _extract_csv(path), 0
    try:
        return Path(path).read_text(encoding="utf-8", errors="replace"), 0
    except Exception:
        return "", 0

def _chunk_text(text: str, size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> List[str]:
    sentences = re.split(r'(?<=[.!?])\s+', text.replace("\n\n", "  "))
    chunks: List[str] = []
    current = ""
    for sent in sentences:
        if len(current) + len(sent) + 1 > size and current:
            chunks.append(current.strip())
            words = current.split()
            carry = " ".join(words[max(0, len(words) - overlap // 6):])
            current = carry + " " + sent
        else:
            current = (current + " " + sent).strip() if current else sent
    if current.strip():
        chunks.append(current.strip())
    return [c for c in chunks if len(c) >= 40]

def _bm25_score(query_terms: List[str], doc: str, avg_len: float = 400, k1: float = 1.5, b: float = 0.75) -> float:
    doc_lower = doc.lower()
    words     = doc_lower.split()
    doc_len   = len(words)
    score     = 0.0
    for term in query_terms:
        tf = doc_lower.count(term)
        if tf == 0:
            continue
        idf     = math.log(1 + 1.0 / (tf + 0.5))
        tf_norm = (tf * (k1 + 1)) / (tf + k1 * (1 - b + b * doc_len / max(avg_len, 1)))
        score  += idf * tf_norm
    return score

async def _rerank(query: str, chunks: List[Dict]) -> List[Dict]:
    if len(chunks) <= RERANK_TOP_K:
        return chunks[:RERANK_TOP_K]
    try:
        numbered = "\n".join(f"[{i}] {c['text'][:200]}" for i, c in enumerate(chunks))
        prompt = (
            f"Query: {query}\n\nScore each passage 0-10 for relevance.\n"
            f"Return ONLY JSON array of scores: {numbered}\n\nFormat: [8, 3, 9, 2, 6, 7]"
        )
        result = await _llm_text(
            [{"role": "user", "content": prompt}], model_id="fast", max_tokens=80, temperature=0.0,
        )
        raw    = re.sub(r"[^\d,\[\] ]", "", result.strip())
        scores = json.loads(raw)
        if isinstance(scores, list) and len(scores) == len(chunks):
            scored = sorted(zip(chunks, scores), key=lambda x: x[1], reverse=True)
            return [c for c, _ in scored[:RERANK_TOP_K]]
    except Exception:
        pass
    return chunks[:RERANK_TOP_K]

async def _rag_retrieve(user_id: str, query: str) -> str:
    if _docs_collection is None:
        return ""
    try:
        results = _docs_collection.query(
            query_texts=[query],
            n_results=min(TOP_K_RETRIEVAL * 2, max(_docs_collection.count(), 1)),
            where={"user_id": user_id},
            include=["documents", "metadatas", "distances"],
        )
        if not results or not results["documents"] or not results["documents"][0]:
            return ""
        docs   = results["documents"][0]
        metas  = results["metadatas"][0] if results.get("metadatas") else [{}] * len(docs)
        dists  = results["distances"][0] if results.get("distances") else [0.5] * len(docs)
        candidates = [
            {"text": d, "meta": m, "semantic_rank": i + 1, "semantic_score": 1 - dist}
            for i, (d, m, dist) in enumerate(zip(docs, metas, dists))
        ]
        query_terms = [t.lower() for t in query.split() if len(t) > 3]
        avg_len = sum(len(c["text"].split()) for c in candidates) / max(len(candidates), 1)
        for c in candidates:
            c["bm25_score"] = _bm25_score(query_terms, c["text"], avg_len)
        bm25_ranked = sorted(candidates, key=lambda x: x["bm25_score"], reverse=True)
        for rank, c in enumerate(bm25_ranked):
            c["bm25_rank"] = rank + 1
        k = 60
        for c in candidates:
            c["rrf_score"] = (1 / (k + c["semantic_rank"])) + (1 / (k + c["bm25_rank"]))
        fused    = sorted(candidates, key=lambda x: x["rrf_score"], reverse=True)[:TOP_K_RETRIEVAL]
        reranked = await _rerank(query, fused)
        parts: List[str] = []
        for c in reranked:
            fname = c["meta"].get("original_name", c["meta"].get("filename", "document"))
            parts.append(f"[Source: {fname}]\n{c['text']}")
        return "\n\n---\n\n".join(parts)
    except Exception as exc:
        logger.error("[RAG] Retrieval failed: %s", exc)
        return ""

async def _ingest_document(doc_id: str, file_path: str, mime: str, user_id: str) -> None:
    t0 = time.time()
    try:
        text, pages = await asyncio.get_event_loop().run_in_executor(
            _executor, _parse_document, file_path, mime
        )
        if not text.strip():
            await db_execute("UPDATE documents SET index_error=? WHERE id=?", ("No text extracted", doc_id))
            return
        chunks = _chunk_text(text)
        if not chunks:
            await db_execute("UPDATE documents SET index_error=? WHERE id=?", ("No chunks produced", doc_id))
            return
        chroma_ids = [f"{doc_id}_c{i}" for i in range(len(chunks))]
        metas = [
            {"user_id": user_id, "doc_id": doc_id, "chunk_index": i, "original_name": Path(file_path).name}
            for i in range(len(chunks))
        ]
        if _docs_collection is not None:
            for batch_start in range(0, len(chunks), 100):
                _docs_collection.upsert(
                    documents=chunks[batch_start:batch_start + 100],
                    ids=chroma_ids[batch_start:batch_start + 100],
                    metadatas=metas[batch_start:batch_start + 100],
                )
        now = _utcnow()
        meta_rows = [
            (_new_id(), doc_id, user_id, i, chunks[i][:500], chroma_ids[i],
             "all-MiniLM-L6-v2", _count_tokens(chunks[i]), now)
            for i in range(len(chunks))
        ]
        await db_executemany(
            "INSERT OR IGNORE INTO embeddings_meta"
            " (id,document_id,user_id,chunk_index,chunk_text,chroma_id,embed_model,token_count,created_at)"
            " VALUES (?,?,?,?,?,?,?,?,?)", meta_rows,
        )
        elapsed = int((time.time() - t0) * 1000)
        await db_execute(
            "UPDATE documents SET is_indexed=1, chunk_count=?, char_count=?, page_count=?, indexed_at=? WHERE id=?",
            (len(chunks), len(text), pages, now, doc_id),
        )
        logger.info("[RAG] Ingested doc %s: %d chunks in %dms", doc_id, len(chunks), elapsed)
    except Exception as exc:
        logger.error("[RAG] Ingest failed for %s: %s", doc_id, exc)
        await db_execute("UPDATE documents SET index_error=? WHERE id=?", (str(exc)[:500], doc_id))


# ══════════════════════════════════════════════════════════════════════════════
# §11  CONTEXT BUILDER
# ══════════════════════════════════════════════════════════════════════════════

async def _get_session_summaries(session_id: str) -> str:
    rows = await db_fetchall(
        "SELECT summary_text FROM summaries WHERE session_id=? ORDER BY created_at", (session_id,),
    )
    if not rows:
        return ""
    return "[Conversation summary (earlier turns)]\n" + "\n---\n".join(r["summary_text"] for r in rows)

async def _build_context(
    session_id: str, user_id: str, user_message: str,
    model_id: str = "censored", use_rag: bool = True,
) -> List[Dict]:
    budget = MAX_CONTEXT_TOKENS
    system_parts: List[str] = [_NORMAL_SYSTEM]
    memories  = await _get_active_memories(user_id, 30)
    mem_block = _format_memories(memories)
    if mem_block:
        system_parts.append(mem_block)
    summary_block = await _get_session_summaries(session_id)
    if summary_block:
        system_parts.append(summary_block)
    rag_context = ""
    if use_rag:
        rag_raw = await _rag_retrieve(user_id, user_message)
        if rag_raw:
            rag_tokens = _count_tokens(rag_raw)
            if rag_tokens <= budget // 3:
                rag_context = f"\n[Retrieved document context:]\n{rag_raw}"
                budget -= rag_tokens
            else:
                max_chars = (budget // 3) * 4
                rag_context = f"\n[Retrieved document context:]\n{rag_raw[:max_chars]}…"
                budget -= budget // 3
    system_content = "\n\n".join(filter(None, system_parts))
    if rag_context:
        system_content += rag_context
    messages: List[Dict] = [{"role": "system", "content": system_content}]
    budget -= _count_tokens(system_content)
    history = await db_fetchall(
        "SELECT role, content FROM chat_history"
        " WHERE session_id=? AND user_id=? AND is_hidden=0"
        " ORDER BY created_at DESC LIMIT 60",
        (session_id, user_id),
    )
    history = list(reversed(history))
    window: List[Dict] = []
    for msg in reversed(history):
        tok = _count_tokens(msg["content"])
        if budget - tok < 200:
            break
        window.insert(0, {"role": msg["role"], "content": msg["content"]})
        budget -= tok
    messages.extend(window)
    messages.append({"role": "user", "content": user_message})
    return messages

async def _maybe_summarize(session_id: str, user_id: str) -> None:
    turn_count = await db_count(
        "SELECT COUNT(*) FROM chat_history WHERE session_id=? AND is_hidden=0", (session_id,),
    )
    if turn_count < SUMMARY_TRIGGER:
        return
    oldest = await db_fetchall(
        "SELECT id, role, content, created_at FROM chat_history"
        " WHERE session_id=? AND is_hidden=0 ORDER BY created_at ASC LIMIT ?",
        (session_id, SUMMARY_TRIGGER),
    )
    if not oldest:
        return
    conv_text = "\n".join(f"{m['role'].capitalize()}: {m['content'][:300]}" for m in oldest)
    prompt = (
        "Compress the following conversation into a concise 3-5 sentence summary.\n"
        "Preserve key facts, decisions, and user preferences.\n\n" + conv_text
    )
    try:
        summary_text = await _llm_text(
            [{"role": "user", "content": prompt}], model_id="fast", max_tokens=250, temperature=0.3,
        )
        covers_from  = oldest[0]["created_at"]
        covers_to    = oldest[-1]["created_at"]
        tokens_saved = sum(_count_tokens(m["content"]) for m in oldest)
        await db_execute(
            "INSERT INTO summaries (id,session_id,user_id,summary_text,covers_from_at,"
            "covers_to_at,turn_count,tokens_saved,model_used,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (_new_id(), session_id, user_id, summary_text, covers_from, covers_to,
             len(oldest), tokens_saved, "fast", _utcnow()),
        )
        msg_ids = tuple(m["id"] for m in oldest)
        placeholders = ",".join("?" * len(msg_ids))
        await db_execute(
            f"UPDATE chat_history SET is_hidden=1 WHERE id IN ({placeholders})", msg_ids
        )
    except Exception as exc:
        logger.warning("[SUMMARY] Failed: %s", exc)


# ══════════════════════════════════════════════════════════════════════════════
# §12  CHAT ENGINE
# ══════════════════════════════════════════════════════════════════════════════

async def _auto_title(session_id: str, user_message: str) -> None:
    try:
        result = await _llm_text(
            [{"role": "user", "content":
              f"Generate a short 4-6 word chat title for: {user_message[:200]}\n"
              "Return ONLY the title — no quotes, no punctuation."}],
            model_id="fast", max_tokens=30, temperature=0.4,
        )
        title = result.strip()[:80]
        if title:
            await db_execute(
                "UPDATE chat_sessions SET title=?, updated_at=? WHERE id=?",
                (title, _utcnow(), session_id),
            )
    except Exception:
        pass

async def _get_or_create_session(user_id: str, session_id: Optional[str], model_id: str) -> str:
    if session_id:
        row = await db_fetchone(
            "SELECT id FROM chat_sessions WHERE id=? AND user_id=?", (session_id, user_id)
        )
        if row:
            return session_id
    sid = _new_id(); now = _utcnow()
    await db_execute(
        "INSERT INTO chat_sessions (id,user_id,title,model_id,created_at,updated_at)"
        " VALUES (?,?,?,?,?,?)",
        (sid, user_id, "New Chat", model_id, now, now),
    )
    return sid

async def _save_turn(
    session_id: str, user_id: str,
    user_msg: str, assistant_msg: str,
    model_id: str, agent_name: str,
    rag_chunks: List[str],
    tok_in: int, tok_out: int, latency_ms: int,
    thinking_content: str = "",
    mode: str = "normal",
) -> Tuple[str, str]:
    now         = _utcnow()
    user_id_msg = _new_id()
    asst_id     = _new_id()
    rag_json    = json.dumps(rag_chunks)
    await db_execute(
        "INSERT INTO chat_history (id,session_id,user_id,role,content,"
        " model_used,agent_used,tool_calls_json,rag_chunks_json,"
        " tokens_input,tokens_output,latency_ms,thinking_content,mode,created_at)"
        " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (user_id_msg, session_id, user_id, "user", user_msg,
         model_id, agent_name, "[]", "[]", tok_in, 0, 0, "", mode, now),
    )
    await db_execute(
        "INSERT INTO chat_history (id,session_id,user_id,role,content,"
        " model_used,agent_used,tool_calls_json,rag_chunks_json,"
        " tokens_input,tokens_output,latency_ms,thinking_content,mode,created_at)"
        " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (asst_id, session_id, user_id, "assistant", assistant_msg,
         model_id, agent_name, "[]", rag_json, tok_in, tok_out, latency_ms, thinking_content, mode, now),
    )
    await db_execute(
        "UPDATE chat_sessions SET turn_count=turn_count+1, last_message_at=?, updated_at=? WHERE id=?",
        (now, now, session_id),
    )
    return user_id_msg, asst_id

async def chat_send_message(
    user_id: str, user_message: str,
    session_id: Optional[str],
    model_id: str = "censored",
    use_rag: bool = True,
    background: Optional[BackgroundTasks] = None,
    force_thinking: Optional[bool] = None,
) -> Dict:
    t0    = time.time()
    sid   = await _get_or_create_session(user_id, session_id, model_id)
    messages = await _build_context(sid, user_id, user_message, model_id, use_rag)
    tok_in   = _msgs_tokens(messages)
    agent    = _route_message(user_message)

    # Auto-detect thinking mode
    use_thinking = force_thinking if force_thinking is not None else _should_think(user_message)
    thinking_content = ""
    mode = "normal"

    if use_thinking:
        mode = "thinking"
        thinking_content, response = await _llm_thinking(messages, model_id, max_tokens=4096)
    else:
        response = await _llm_text(messages, model_id, max_tokens=2048, temperature=0.7)

    tok_out  = _count_tokens(response)
    elapsed  = int((time.time() - t0) * 1000)

    user_msg_id, asst_id = await _save_turn(
        sid, user_id, user_message, response, model_id,
        agent["name"], [], tok_in, tok_out, elapsed,
        thinking_content=thinking_content, mode=mode,
    )

    session_row = await db_fetchone("SELECT turn_count FROM chat_sessions WHERE id=?", (sid,))
    if session_row and session_row["turn_count"] == 1:
        if background:
            background.add_task(_auto_title, sid, user_message)

    if background:
        background.add_task(_extract_and_save_memories, user_id, user_message, response)
        background.add_task(_maybe_summarize, sid, user_id)

    return {
        "session_id": sid, "message_id": asst_id,
        "response": response, "agent": agent["name"], "model": model_id,
        "tokens": {"input": tok_in, "output": tok_out},
        "latency_ms": elapsed, "mode": mode,
        "thinking": thinking_content,
    }

async def chat_stream_message(
    user_id: str, user_message: str,
    session_id: Optional[str],
    model_id: str = "censored",
    use_rag: bool = True,
    force_thinking: Optional[bool] = None,
) -> AsyncGenerator[str, None]:
    t0    = time.time()
    sid   = await _get_or_create_session(user_id, session_id, model_id)
    agent = _route_message(user_message)

    # Auto-detect thinking
    use_thinking = force_thinking if force_thinking is not None else _should_think(user_message)
    mode = "thinking" if use_thinking else "normal"

    yield f"data: {json.dumps({'type':'start','session_id':sid,'agent':agent['name'],'mode':mode})}\n\n"

    # For thinking mode, do non-streaming with thinking extraction
    if use_thinking:
        try:
            messages = await _build_context(sid, user_id, user_message, model_id, use_rag)
            tok_in   = _msgs_tokens(messages)
            yield f"data: {json.dumps({'type':'thinking_start'})}\n\n"
            thinking_content, response = await _llm_thinking(messages, model_id, max_tokens=4096)
            yield f"data: {json.dumps({'type':'thinking_content','content':thinking_content})}\n\n"
            yield f"data: {json.dumps({'type':'thinking_done'})}\n\n"
            # Stream response char by char
            for i in range(0, len(response), 8):
                chunk = response[i:i+8]
                yield f"data: {json.dumps({'type':'delta','content':chunk})}\n\n"
                await asyncio.sleep(0.01)
            tok_out = _count_tokens(response)
            elapsed = int((time.time() - t0) * 1000)
            _, asst_id = await _save_turn(
                sid, user_id, user_message, response, model_id,
                agent["name"], [], tok_in, tok_out, elapsed,
                thinking_content=thinking_content, mode=mode,
            )
            yield f"data: {json.dumps({'type':'done','message_id':asst_id,'latency_ms':elapsed,'tokens':{'input':tok_in,'output':tok_out},'mode':mode})}\n\n"
            asyncio.create_task(_extract_and_save_memories(user_id, user_message, response))
            asyncio.create_task(_maybe_summarize(sid, user_id))
            return
        except Exception as exc:
            yield f"data: {json.dumps({'type':'error','message':str(exc)})}\n\n"
            return

    # Normal streaming
    messages = await _build_context(sid, user_id, user_message, model_id, use_rag)
    tok_in   = _msgs_tokens(messages)
    loop     = asyncio.get_event_loop()
    chain    = [model_id] + [m for m in _FALLBACK_CHAIN if m != model_id]
    stream_obj = None
    used_model = model_id
    for mid in chain:
        try:
            client, model_name = _get_llm_client(mid)
            stream_obj = await loop.run_in_executor(
                _executor,
                lambda c=client, m=model_name: c.chat.completions.create(
                    model=m, messages=messages, max_tokens=2048, temperature=0.7, stream=True,
                ),
            )
            used_model = mid
            break
        except Exception as exc:
            logger.warning("[STREAM] %s failed: %s", mid, exc)

    if stream_obj is None:
        yield f"data: {json.dumps({'type':'error','message':'All models unavailable'})}\n\n"
        return

    queue: asyncio.Queue = asyncio.Queue()
    _DONE = object()

    def _drain_stream() -> None:
        try:
            for chunk in stream_obj:
                delta = ""
                if chunk.choices and chunk.choices[0].delta:
                    delta = chunk.choices[0].delta.content or ""
                loop.call_soon_threadsafe(queue.put_nowait, delta)
        except Exception as exc:
            loop.call_soon_threadsafe(queue.put_nowait, exc)
        finally:
            loop.call_soon_threadsafe(queue.put_nowait, _DONE)

    loop.run_in_executor(_executor, _drain_stream)

    full_response = ""
    try:
        while True:
            item = await asyncio.wait_for(queue.get(), timeout=60.0)
            if item is _DONE:
                break
            if isinstance(item, Exception):
                yield f"data: {json.dumps({'type':'error','message':str(item)})}\n\n"
                return
            if item:
                full_response += item
                yield f"data: {json.dumps({'type':'delta','content':item})}\n\n"
    except asyncio.TimeoutError:
        yield f"data: {json.dumps({'type':'error','message':'Stream timeout'})}\n\n"
        return

    tok_out = _count_tokens(full_response)
    elapsed = int((time.time() - t0) * 1000)
    _, asst_id = await _save_turn(
        sid, user_id, user_message, full_response, used_model,
        agent["name"], [], tok_in, tok_out, elapsed, mode="normal",
    )
    yield f"data: {json.dumps({'type':'done','message_id':asst_id,'latency_ms':elapsed,'tokens':{'input':tok_in,'output':tok_out},'mode':'normal'})}\n\n"
    asyncio.create_task(_extract_and_save_memories(user_id, user_message, full_response))
    asyncio.create_task(_maybe_summarize(sid, user_id))


# ══════════════════════════════════════════════════════════════════════════════
# §13  AGENT ROUTER
# ══════════════════════════════════════════════════════════════════════════════

_AGENTS: List[Dict] = [
    {"name": "coding_agent",   "keywords": ["code","python","script","function","class","debug","algorithm","api","bug","javascript","typescript","sql","refactor","implement","fix","lint","build","deploy"]},
    {"name": "research_agent", "keywords": ["research","explain","analyze","investigate","compare","review","report","summarize","difference","how does","why does","what is","overview","describe"]},
    {"name": "data_agent",     "keywords": ["data","chart","plot","statistics","average","trend","insight","metrics","kpi","dashboard","forecast","csv","excel"]},
    {"name": "vision_agent",   "keywords": ["image","photo","picture","screenshot","diagram","see","look","visual","analyze image"]},
    {"name": "system_agent",   "keywords": ["run","execute","open","launch","install","computer","system","terminal","shell","cmd","file","folder","directory","process","browser","camera","mic","screen"]},
    {"name": "general_agent",  "keywords": []},
]

def _route_message(message: str, has_image: bool = False) -> Dict:
    if has_image:
        return next(a for a in _AGENTS if a["name"] == "vision_agent")
    low   = message.lower()
    best  = _AGENTS[-1]
    best_score = 0
    for agent in _AGENTS[:-1]:
        score = sum(1 for kw in agent["keywords"] if kw in low)
        if score > best_score:
            best, best_score = agent, score
    return best


# ══════════════════════════════════════════════════════════════════════════════
# §14  COMPUTER / SYSTEM ACCESS
# ══════════════════════════════════════════════════════════════════════════════

class ToolResult(BaseModel):
    success: bool
    output:  Any
    error:   Optional[str] = None


def _run_shell(command: str, timeout: int = CODE_EXEC_TIMEOUT, cwd: Optional[str] = None) -> Dict:
    """Run any shell command — full system access."""
    t0 = time.time()
    try:
        result = subprocess.run(
            command, shell=True, capture_output=True, text=True,
            timeout=timeout, cwd=cwd or os.getcwd(),
            env={**os.environ},
        )
        return {
            "stdout":    result.stdout[:CODE_EXEC_MAX_OUT],
            "stderr":    result.stderr[:CODE_EXEC_MAX_OUT],
            "exit_code": result.returncode,
            "duration_ms": int((time.time() - t0) * 1000),
        }
    except subprocess.TimeoutExpired:
        return {"stdout": "", "stderr": f"Timed out ({timeout}s)", "exit_code": -1, "duration_ms": timeout * 1000}
    except Exception as exc:
        return {"stdout": "", "stderr": str(exc), "exit_code": -1, "duration_ms": 0}


def _run_code_sync(code: str, language: str, cwd: Optional[str] = None) -> Dict:
    """Run code in any language."""
    t0 = time.time()
    cmd_map = {
        "python":     [sys.executable, "-c", code],
        "javascript": ["node", "-e", code],
        "bash":       ["bash", "-c", code],
        "shell":      ["bash", "-c", code],
        "powershell": ["powershell", "-Command", code],
        "ruby":       ["ruby", "-e", code],
        "php":        ["php", "-r", code],
    }
    cmd = cmd_map.get(language)
    if not cmd:
        # Try as shell command directly
        return _run_shell(code, cwd=cwd)
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True,
            timeout=CODE_EXEC_TIMEOUT, cwd=cwd or os.getcwd(),
            env={**os.environ},
        )
        return {
            "stdout":    result.stdout[:CODE_EXEC_MAX_OUT],
            "stderr":    result.stderr[:CODE_EXEC_MAX_OUT],
            "exit_code": result.returncode,
            "duration_ms": int((time.time() - t0) * 1000),
        }
    except subprocess.TimeoutExpired:
        return {"stdout": "", "stderr": f"Timed out ({CODE_EXEC_TIMEOUT}s)", "exit_code": -1, "duration_ms": CODE_EXEC_TIMEOUT * 1000}
    except FileNotFoundError as exc:
        return {"stdout": "", "stderr": f"Runtime not found: {exc}", "exit_code": -1, "duration_ms": 0}
    except Exception as exc:
        return {"stdout": "", "stderr": str(exc), "exit_code": -1, "duration_ms": 0}


def _open_application(app_name: str) -> Dict:
    """Open application / URL on host system."""
    system = platform.system()
    try:
        if system == "Windows":
            os.startfile(app_name)
        elif system == "Darwin":
            subprocess.Popen(["open", app_name])
        else:
            subprocess.Popen(["xdg-open", app_name])
        return {"success": True, "message": f"Opened: {app_name}"}
    except Exception as exc:
        return {"success": False, "error": str(exc)}


def _list_files(path: str = ".") -> Dict:
    """List files in a directory."""
    try:
        p = Path(path).expanduser().resolve()
        items = []
        for item in p.iterdir():
            try:
                stat = item.stat()
                items.append({
                    "name": item.name,
                    "type": "dir" if item.is_dir() else "file",
                    "size": stat.st_size if item.is_file() else None,
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "path": str(item),
                })
            except Exception:
                pass
        return {"path": str(p), "items": sorted(items, key=lambda x: (x["type"] != "dir", x["name"]))}
    except Exception as exc:
        return {"error": str(exc)}


def _read_file(path: str, max_bytes: int = 100_000) -> Dict:
    """Read file contents."""
    try:
        p = Path(path).expanduser().resolve()
        if not p.exists():
            return {"error": f"File not found: {path}"}
        if p.stat().st_size > max_bytes * 10:
            return {"error": "File too large (> 1MB)"}
        content = p.read_bytes()[:max_bytes]
        try:
            return {"path": str(p), "content": content.decode("utf-8", errors="replace"), "size": p.stat().st_size}
        except Exception:
            return {"path": str(p), "content": f"[Binary file, {p.stat().st_size} bytes]", "size": p.stat().st_size}
    except Exception as exc:
        return {"error": str(exc)}


def _write_file(path: str, content: str) -> Dict:
    """Write content to file."""
    try:
        p = Path(path).expanduser().resolve()
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content, encoding="utf-8")
        return {"success": True, "path": str(p), "size": p.stat().st_size}
    except Exception as exc:
        return {"error": str(exc)}


def _get_system_info() -> Dict:
    """Get system information."""
    try:
        return {
            "os": platform.system(),
            "os_version": platform.version(),
            "machine": platform.machine(),
            "processor": platform.processor(),
            "python": sys.version,
            "cwd": os.getcwd(),
            "home": str(Path.home()),
            "hostname": platform.node(),
        }
    except Exception as exc:
        return {"error": str(exc)}


# ══════════════════════════════════════════════════════════════════════════════
# §15  TOOL REGISTRY
# ══════════════════════════════════════════════════════════════════════════════

async def _tool_code_exec(args: Dict) -> ToolResult:
    code     = args.get("code", "")
    language = args.get("language", "python")
    cwd      = args.get("cwd")
    result = await asyncio.get_event_loop().run_in_executor(
        _executor, lambda: _run_code_sync(code, language, cwd)
    )
    return ToolResult(
        success=result["exit_code"] == 0,
        output=result,
        error=result["stderr"] if result["exit_code"] != 0 else None,
    )

async def _tool_shell(args: Dict) -> ToolResult:
    command = args.get("command", "")
    timeout = args.get("timeout", CODE_EXEC_TIMEOUT)
    cwd     = args.get("cwd")
    result  = await asyncio.get_event_loop().run_in_executor(
        _executor, lambda: _run_shell(command, timeout, cwd)
    )
    return ToolResult(
        success=result["exit_code"] == 0,
        output=result,
        error=result["stderr"] if result["exit_code"] != 0 else None,
    )

async def _tool_file_ops(args: Dict) -> ToolResult:
    op   = args.get("op", "read")
    path = args.get("path", ".")
    if op == "list":
        result = await asyncio.get_event_loop().run_in_executor(_executor, lambda: _list_files(path))
        return ToolResult(success="error" not in result, output=result, error=result.get("error"))
    elif op == "read":
        result = await asyncio.get_event_loop().run_in_executor(_executor, lambda: _read_file(path))
        return ToolResult(success="error" not in result, output=result, error=result.get("error"))
    elif op == "write":
        content = args.get("content", "")
        result  = await asyncio.get_event_loop().run_in_executor(_executor, lambda: _write_file(path, content))
        return ToolResult(success=result.get("success", False), output=result, error=result.get("error"))
    elif op == "delete":
        try:
            p = Path(path).expanduser().resolve()
            if p.is_dir():
                shutil.rmtree(p)
            else:
                p.unlink()
            return ToolResult(success=True, output={"deleted": str(p)})
        except Exception as exc:
            return ToolResult(success=False, output=None, error=str(exc))
    return ToolResult(success=False, output=None, error=f"Unknown op: {op}")

async def _tool_open_app(args: Dict) -> ToolResult:
    app = args.get("app", "")
    result = await asyncio.get_event_loop().run_in_executor(_executor, lambda: _open_application(app))
    return ToolResult(success=result.get("success", False), output=result, error=result.get("error"))

async def _tool_system_info(args: Dict) -> ToolResult:
    result = await asyncio.get_event_loop().run_in_executor(_executor, _get_system_info)
    return ToolResult(success=True, output=result)

async def _tool_db_query(args: Dict) -> ToolResult:
    question = args.get("question", "")
    try:
        sql_prompt = (
            f"Write a safe SELECT SQL for SQLite: {question}\n"
            "Rules: SELECT only. Add LIMIT 100. Return ONLY SQL."
        )
        sql = await _llm_text([{"role": "user", "content": sql_prompt}], model_id="fast", max_tokens=200)
        sql = re.sub(r"```sql|```", "", sql).strip()
        if not sql.upper().startswith("SELECT"):
            return ToolResult(success=False, output=None, error="Only SELECT allowed")
        rows = await db_fetchall(sql)
        return ToolResult(success=True, output={"sql": sql, "rows": rows, "count": len(rows)})
    except Exception as exc:
        return ToolResult(success=False, output=None, error=str(exc))

async def _tool_rag_search(args: Dict, user_id: str) -> ToolResult:
    query  = args.get("query", "")
    result = await _rag_retrieve(user_id, query)
    return ToolResult(success=bool(result), output=result or "No results found")

async def _tool_connector_action(args: Dict, user_id: str) -> ToolResult:
    connector_name = args.get("connector")
    action         = args.get("action")
    params         = args.get("params", {})
    row = await db_fetchone(
        "SELECT * FROM connectors WHERE user_id=? AND name=? AND is_active=1", (user_id, connector_name),
    )
    if not row:
        return ToolResult(success=False, output=None, error=f"Connector '{connector_name}' not found")
    try:
        creds  = _decrypt(row["encrypted_creds"])
        result = await _connector_dispatch(row["connector_type"], action, creds, params)
        return ToolResult(success=True, output=result)
    except Exception as exc:
        return ToolResult(success=False, output=None, error=str(exc))

_TOOL_REGISTRY: Dict[str, Any] = {
    "code_exec":        _tool_code_exec,
    "shell":            _tool_shell,
    "file_ops":         _tool_file_ops,
    "open_app":         _tool_open_app,
    "system_info":      _tool_system_info,
    "db_query":         _tool_db_query,
    "rag_search":       _tool_rag_search,
    "connector_action": _tool_connector_action,
}

async def _dispatch_tool(tool_name: str, args: Dict, user_id: str = "") -> ToolResult:
    fn = _TOOL_REGISTRY.get(tool_name)
    if fn is None:
        return ToolResult(success=False, output=None, error=f"Unknown tool: {tool_name}")
    try:
        import inspect
        sig = inspect.signature(fn)
        if "user_id" in sig.parameters:
            return await fn(args, user_id)
        return await fn(args)
    except Exception as exc:
        return ToolResult(success=False, output=None, error=str(exc))


# ══════════════════════════════════════════════════════════════════════════════
# §16  AGENT ENGINE
# ══════════════════════════════════════════════════════════════════════════════

_AGENT_SYSTEM = """You are a reasoning agent with FULL system access. Work through tasks step by step.

Available tools:
  code_exec(code, language, cwd?)      — run Python/JS/bash/shell code
  shell(command, timeout?, cwd?)       — execute any shell command
  file_ops(op, path, content?)         — op: list|read|write|delete
  open_app(app)                        — open application, URL, or file
  system_info()                        — get system/OS information
  db_query(question)                   — query database in natural language
  rag_search(query)                    — search user documents
  connector_action(connector, action, params) — call external connector

On each step respond ONLY with valid JSON:
{
  "thought": "what I'm thinking",
  "action": "tool_name | FINISH",
  "args": {},
  "final_answer": "only when action=FINISH"
}
Max 12 steps. You have full access to the computer/OS/files."""

_MAX_AGENT_STEPS = 12

async def _run_agent_loop(
    user_id: str, prompt: str, allowed_tools: List[str],
    model_id: str = "censored", timeout_s: int = 120,
) -> Tuple[str, List[Dict]]:
    messages: List[Dict] = [
        {"role": "system", "content": _AGENT_SYSTEM},
        {"role": "user",   "content": prompt},
    ]
    tool_log: List[Dict] = []
    deadline = time.time() + timeout_s

    for step in range(_MAX_AGENT_STEPS):
        if time.time() > deadline:
            break
        try:
            raw = await _llm_text(messages, model_id, max_tokens=800, temperature=0.2)
            clean = re.sub(r"```json|```", "", raw.strip()).strip()
            step_json = json.loads(clean)
        except Exception as exc:
            logger.warning("[AGENT] Step %d parse error: %s", step, exc)
            break

        thought = step_json.get("thought", "")
        action  = step_json.get("action", "FINISH")
        args    = step_json.get("args", {})
        messages.append({"role": "assistant", "content": raw})

        if action == "FINISH":
            return step_json.get("final_answer", "Task complete."), tool_log

        # Allow all tools if list is empty, else filter
        if allowed_tools and action not in allowed_tools:
            obs = f"Tool '{action}' not in allowed list: {allowed_tools}"
        else:
            t_start = time.time()
            result  = await _dispatch_tool(action, args, user_id)
            obs = json.dumps({
                "success": result.success,
                "output":  str(result.output)[:3000] if result.output else None,
                "error":   result.error,
            })
            tool_log.append({
                "step": step, "tool": action, "args": args,
                "success": result.success, "latency_ms": int((time.time() - t_start) * 1000),
            })

        messages.append({"role": "user", "content": f"[Observation]: {obs}"})

    return "Agent could not complete the task within the step limit.", tool_log


# ══════════════════════════════════════════════════════════════════════════════
# §17  BACKGROUND JOB SCHEDULER
# ══════════════════════════════════════════════════════════════════════════════

_scheduler: Optional[BackgroundScheduler] = None

async def _run_job(job_id: str) -> None:
    row = await db_fetchone("SELECT * FROM agent_jobs WHERE id=? AND enabled=1", (job_id,))
    if not row or row["status"] == "running":
        return
    run_num = (row.get("total_runs") or 0) + 1
    log_id  = _new_id()
    started = _utcnow()
    await db_execute("UPDATE agent_jobs SET status='running', updated_at=? WHERE id=?", (_utcnow(), job_id))
    await db_execute(
        "INSERT INTO agent_job_logs (id,job_id,user_id,run_number,status,prompt_rendered,started_at)"
        " VALUES (?,?,?,?,?,?,?)",
        (log_id, job_id, row["user_id"], run_num, "running", row["prompt_template"], started),
    )
    tools       = json.loads(row.get("tools_json", "[]"))
    max_ret     = row.get("max_retries", 2)
    result_text = None
    tool_log    = []
    run_status  = "failure"
    error_msg   = None
    t0          = time.time()
    for attempt in range(max_ret + 1):
        try:
            result_text, tool_log = await _run_agent_loop(
                user_id=row["user_id"], prompt=row["prompt_template"],
                allowed_tools=tools, model_id=row.get("model_id", "censored"),
                timeout_s=row.get("timeout_seconds", 120),
            )
            run_status = "success"
            error_msg  = None
            break
        except Exception as exc:
            error_msg = str(exc)
            await asyncio.sleep(min(2 ** attempt, 30))
    elapsed = int((time.time() - t0) * 1000)
    now     = _utcnow()
    await db_execute(
        "UPDATE agent_job_logs SET status=?, result_text=?, tool_calls_json=?,"
        " latency_ms=?, error_message=?, finished_at=? WHERE id=?",
        (run_status, result_text, json.dumps(tool_log), elapsed, error_msg, now, log_id),
    )
    col = "success_runs=success_runs+1" if run_status == "success" else "failed_runs=failed_runs+1"
    await db_execute(
        f"UPDATE agent_jobs SET status='idle', total_runs=total_runs+1, {col},"
        " last_run_at=?, last_run_status=?, retry_count=0, updated_at=? WHERE id=?",
        (now, run_status, now, job_id),
    )

def _sync_run_job(job_id: str) -> None:
    asyncio.run(_run_job(job_id))

def _schedule_job(job: Dict) -> None:
    if _scheduler is None:
        return
    jid        = job["id"]
    jtype      = job.get("job_type", "cron")
    trigger_cfg = json.loads(job.get("trigger_json", "{}"))
    job_kwargs  = {"func": _sync_run_job, "args": [jid], "id": jid, "replace_existing": True}
    if jtype == "cron":
        expr   = trigger_cfg.get("cron", "0 9 * * *")
        parts  = expr.split()
        if len(parts) >= 5:
            job_kwargs["trigger"] = CronTrigger(
                minute=parts[0], hour=parts[1], day=parts[2], month=parts[3], day_of_week=parts[4],
            )
        else:
            job_kwargs["trigger"] = IntervalTrigger(hours=24)
    elif jtype == "manual":
        return
    else:
        interval = trigger_cfg.get("interval_seconds", 3600)
        job_kwargs["trigger"] = IntervalTrigger(seconds=interval)
    _scheduler.add_job(**job_kwargs)

def _start_scheduler() -> None:
    global _scheduler
    _scheduler = BackgroundScheduler(timezone="UTC")
    _scheduler.start()

async def _reload_all_jobs() -> None:
    rows = await db_fetchall("SELECT * FROM agent_jobs WHERE enabled=1")
    for row in rows:
        _schedule_job(row)


# ══════════════════════════════════════════════════════════════════════════════
# §18  CONNECTOR SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

async def _connector_dispatch(connector_type: str, action: str, creds: Dict, params: Dict) -> Any:
    handlers = {
        "slack":         _connector_slack,
        "google_sheets": _connector_gsheets,
        "excel":         _connector_excel,
        "power_bi":      _connector_power_bi,
        "http":          _connector_http,
    }
    fn = handlers.get(connector_type)
    if fn is None:
        raise ValueError(f"Unknown connector type: {connector_type}")
    return await fn(action, creds, params)

async def _connector_slack(action: str, creds: Dict, params: Dict) -> Any:
    import urllib.request
    token   = creds.get("bot_token", SLACK_BOT_TOKEN)
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"}
    def _post(url, body): req = urllib.request.Request(url, json.dumps(body).encode(), headers); resp = urllib.request.urlopen(req, timeout=15); return json.loads(resp.read())
    def _get(url): req = urllib.request.Request(url, headers=headers); resp = urllib.request.urlopen(req, timeout=15); return json.loads(resp.read())
    loop = asyncio.get_event_loop()
    if action == "send_message":
        body: Dict = {"channel": params.get("channel", "#general"), "text": params.get("text", "")}
        if params.get("blocks"): body["blocks"] = params["blocks"]
        result = await loop.run_in_executor(_executor, _post, "https://slack.com/api/chat.postMessage", body)
        if not result.get("ok"): raise ValueError(f"Slack error: {result.get('error')}")
        return {"ts": result.get("ts"), "channel": result.get("channel")}
    elif action == "get_messages":
        result = await loop.run_in_executor(_executor, _get, f"https://slack.com/api/conversations.history?channel={params.get('channel_id')}&limit={params.get('limit',20)}")
        if not result.get("ok"): raise ValueError(f"Slack error: {result.get('error')}")
        return result.get("messages", [])
    elif action == "list_channels":
        result = await loop.run_in_executor(_executor, _get, "https://slack.com/api/conversations.list?limit=100")
        return [{"id": c["id"], "name": c["name"]} for c in result.get("channels", [])]
    raise ValueError(f"Unknown Slack action: {action}")

async def _connector_gsheets(action: str, creds: Dict, params: Dict) -> Any:
    import urllib.request, urllib.parse
    token    = creds.get("access_token", "")
    sheet_id = params.get("spreadsheet_id", creds.get("spreadsheet_id", ""))
    headers  = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    def _get(url): req = urllib.request.Request(url, headers=headers); resp = urllib.request.urlopen(req, timeout=20); return json.loads(resp.read())
    def _put(url, body): data = json.dumps(body).encode(); req = urllib.request.Request(url, data, headers, method="PUT"); resp = urllib.request.urlopen(req, timeout=20); return json.loads(resp.read())
    loop = asyncio.get_event_loop()
    base = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}"
    if action == "read_range":
        rng    = urllib.parse.quote(params.get("range", "Sheet1!A1:Z100"))
        result = await loop.run_in_executor(_executor, _get, f"{base}/values/{rng}")
        return result.get("values", [])
    elif action == "write_range":
        rng  = urllib.parse.quote(params.get("range", "Sheet1!A1"))
        body = {"values": params.get("values", []), "majorDimension": "ROWS"}
        return await loop.run_in_executor(_executor, _put, f"{base}/values/{rng}?valueInputOption=USER_ENTERED", body)
    elif action == "get_metadata":
        result = await loop.run_in_executor(_executor, _get, base)
        return {"title": result.get("properties", {}).get("title"), "sheets": [s["properties"]["title"] for s in result.get("sheets", [])]}
    raise ValueError(f"Unknown Google Sheets action: {action}")

async def _connector_excel(action: str, creds: Dict, params: Dict) -> Any:
    file_path = creds.get("file_path", params.get("file_path", ""))
    if not HAS_OPENPYXL:
        raise ValueError("openpyxl not installed")
    def _read():
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[params.get("sheet", wb.sheetnames[0])]
        return {"sheet": ws.title, "rows": [[cell.value for cell in row] for row in ws.iter_rows()]}
    def _write():
        wb = openpyxl.load_workbook(file_path) if Path(file_path).exists() else openpyxl.Workbook()
        sheet_name = params.get("sheet", wb.sheetnames[0] if wb.sheetnames else "Sheet1")
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        for r_idx, row_data in enumerate(params.get("rows", []), start=params.get("start_row", 1)):
            for c_idx, val in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        wb.save(file_path); return {"written_rows": len(params.get("rows", []))}
    loop = asyncio.get_event_loop()
    if action == "read": return await loop.run_in_executor(_executor, _read)
    elif action == "write": return await loop.run_in_executor(_executor, _write)
    raise ValueError(f"Unknown Excel action: {action}")

async def _connector_power_bi(action: str, creds: Dict, params: Dict) -> Any:
    import urllib.request
    token        = creds.get("access_token", "")
    headers      = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    workspace_id = creds.get("workspace_id", params.get("workspace_id", ""))
    dataset_id   = params.get("dataset_id", creds.get("dataset_id", ""))
    base_url     = "https://api.powerbi.com/v1.0/myorg"
    def _get(url): req = urllib.request.Request(url, headers=headers); resp = urllib.request.urlopen(req, timeout=20); return json.loads(resp.read())
    def _post(url, body): req = urllib.request.Request(url, json.dumps(body).encode(), headers); resp = urllib.request.urlopen(req, timeout=30); return json.loads(resp.read())
    loop = asyncio.get_event_loop()
    if action == "list_datasets":
        result = await loop.run_in_executor(_executor, _get, f"{base_url}/groups/{workspace_id}/datasets")
        return [{"id": d["id"], "name": d["name"]} for d in result.get("value", [])]
    elif action == "execute_query":
        url    = f"{base_url}/groups/{workspace_id}/datasets/{dataset_id}/executeQueries"
        body   = {"queries": [{"query": params.get("query", "")}], "serializerSettings": {"includeNulls": True}}
        result = await loop.run_in_executor(_executor, _post, url, body)
        return result.get("results", [{}])[0].get("tables", [])
    raise ValueError(f"Unknown Power BI action: {action}")

async def _connector_http(action: str, creds: Dict, params: Dict) -> Any:
    import urllib.request, urllib.parse
    base_url  = creds.get("base_url", "").rstrip("/")
    auth_type = creds.get("auth_type", "bearer")
    auth_value = creds.get("auth_value", "")
    headers: Dict[str, str] = {"Content-Type": "application/json", **creds.get("headers", {})}
    if auth_type == "bearer": headers["Authorization"] = f"Bearer {auth_value}"
    elif auth_type == "basic": headers["Authorization"] = f"Basic {base64.b64encode(auth_value.encode()).decode()}"
    elif auth_type == "api_key": headers[creds.get("api_key_header", "X-API-Key")] = auth_value
    method  = (params.get("method", action) or "GET").upper()
    path    = params.get("path", "").lstrip("/")
    body    = params.get("body")
    url_params = params.get("params", {})
    full_url = f"{base_url}/{path}"
    if url_params: full_url += "?" + urllib.parse.urlencode(url_params)
    def _call():
        data = json.dumps(body).encode() if body else None
        req  = urllib.request.Request(full_url, data, headers, method=method)
        resp = urllib.request.urlopen(req, timeout=30)
        raw  = resp.read()
        try: return json.loads(raw)
        except: return {"raw": raw.decode("utf-8", errors="replace")}
    return await asyncio.get_event_loop().run_in_executor(_executor, _call)


# ══════════════════════════════════════════════════════════════════════════════
# §19  LIVEKIT VOICE SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

def _lk_available() -> bool:
    return bool(LIVEKIT_API_KEY and LIVEKIT_API_SECRET and LIVEKIT_URL)

def _lk_token(identity: str, name: str, room: str, ttl_s: int) -> str:
    try:
        from livekit.api import AccessToken, VideoGrants
        token = (
            AccessToken(LIVEKIT_API_KEY, LIVEKIT_API_SECRET)
            .with_identity(identity)
            .with_name(name)
            .with_grants(VideoGrants(
                room_join=True, room=room,
                can_publish=True, can_subscribe=True, can_publish_data=True,
            ))
            .with_ttl(timedelta(seconds=ttl_s))
            .to_jwt()
        )
        return token
    except ImportError:
        import hmac, hashlib as _hl
        now     = int(time.time())
        payload = {
            "iss": LIVEKIT_API_KEY, "sub": identity,
            "iat": now, "nbf": now, "exp": now + ttl_s,
            "video": {"roomJoin": True, "room": room, "canPublish": True, "canSubscribe": True, "canPublishData": True},
        }
        header  = base64.urlsafe_b64encode(json.dumps({"alg":"HS256","typ":"JWT"}).encode()).rstrip(b"=").decode()
        body_b  = base64.urlsafe_b64encode(json.dumps(payload).encode()).rstrip(b"=").decode()
        sig_in  = f"{header}.{body_b}".encode()
        sig     = hmac.new(LIVEKIT_API_SECRET.encode(), sig_in, _hl.sha256).digest()
        sig_b64 = base64.urlsafe_b64encode(sig).rstrip(b"=").decode()
        return f"{header}.{body_b}.{sig_b64}"

async def _stt_transcribe(audio_bytes: bytes, mime_type: str = "audio/webm") -> str:
    from openai import OpenAI as _OAI
    client = _OAI(base_url="https://api.groq.com/openai/v1", api_key=GROQ_API_KEY)
    ext_map = {
        "audio/webm": "webm", "audio/mp4": "mp4", "audio/wav": "wav",
        "audio/x-wav": "wav", "audio/ogg": "ogg", "audio/mpeg": "mp3",
        "audio/m4a": "m4a", "video/webm": "webm",
    }
    ext = ext_map.get(mime_type, "webm")
    try:
        audio_io = io.BytesIO(audio_bytes)
        audio_io.name = f"audio.{ext}"
        result = await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: client.audio.transcriptions.create(
                model="whisper-large-v3-turbo", file=audio_io,
                response_format="text",  # auto-detect language
            ),
        )
        transcript = str(result).strip()
        logger.info("[STT] Transcribed %d bytes → %d chars", len(audio_bytes), len(transcript))
        return transcript
    except Exception as exc:
        logger.error("[STT] Failed: %s", exc)
        raise HTTPException(500, f"Transcription failed: {exc}")

async def _tts_synthesize(text: str, voice: str = TTS_VOICE) -> bytes:
    from openai import OpenAI as _OAI
    client = _OAI(base_url="https://api.groq.com/openai/v1", api_key=GROQ_API_KEY)
    try:
        response = await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: client.audio.speech.create(
                model="playai-tts", voice=voice, input=text[:4096], response_format="mp3",
            ),
        )
        return response.content
    except Exception as exc:
        logger.error("[TTS] Failed: %s", exc)
        raise HTTPException(500, f"TTS failed: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# §20  WEBSOCKET MANAGER
# ══════════════════════════════════════════════════════════════════════════════

class ConnectionManager:
    def __init__(self):
        self._connections: Dict[str, Set[WebSocket]] = defaultdict(set)

    async def connect(self, user_id: str, ws: WebSocket) -> None:
        await ws.accept()
        self._connections[user_id].add(ws)

    def disconnect(self, user_id: str, ws: WebSocket) -> None:
        self._connections[user_id].discard(ws)
        if not self._connections[user_id]:
            del self._connections[user_id]

    async def send_to_user(self, user_id: str, data: Dict) -> None:
        dead: List[WebSocket] = []
        for ws in list(self._connections.get(user_id, [])):
            try:
                await ws.send_json(data)
            except Exception:
                dead.append(ws)
        for ws in dead:
            self.disconnect(user_id, ws)

    async def broadcast(self, data: Dict, exclude: Optional[str] = None) -> None:
        for uid in list(self._connections.keys()):
            if uid != exclude:
                await self.send_to_user(uid, data)

    def online_users(self) -> List[str]:
        return list(self._connections.keys())

ws_manager = ConnectionManager()


# ══════════════════════════════════════════════════════════════════════════════
# §21  NOTIFICATION
# ══════════════════════════════════════════════════════════════════════════════

async def _push_notification(
    user_id: Optional[str], title: str, message: str,
    ntype: str = "info", link: Optional[str] = None,
    is_broadcast: bool = False, actor_id: Optional[str] = None,
) -> None:
    try:
        nid = _new_id()
        await db_execute(
            "INSERT INTO notifications (id,user_id,title,message,type,link,is_broadcast,actor_id,created_at)"
            " VALUES (?,?,?,?,?,?,?,?,?)",
            (nid, user_id, title, message, ntype, link, int(is_broadcast), actor_id, _utcnow()),
        )
        payload = {
            "type": "notification", "id": nid, "title": title,
            "message": message, "ntype": ntype, "link": link, "created_at": _utcnow(),
        }
        if is_broadcast:
            await ws_manager.broadcast(payload)
        elif user_id:
            await ws_manager.send_to_user(user_id, payload)
    except Exception as exc:
        logger.warning("[NOTIF] %s", exc)


# ══════════════════════════════════════════════════════════════════════════════
# §22  VISION
# ══════════════════════════════════════════════════════════════════════════════

async def _analyze_image(image_data: str, user_message: str, user_id: str) -> Tuple[str, str]:
    vc, vm = _get_vision_client()
    vision_prompt = (
        'Analyze this image, return JSON: '
        '{"description":"...","objects":[],"text_detected":null,'
        '"scene_type":"photo|diagram|screenshot|chart|art|other"}'
    )
    vision_json: Dict = {}
    try:
        vresp = await asyncio.get_event_loop().run_in_executor(
            _executor,
            lambda: vc.chat.completions.create(
                model=vm,
                messages=[{"role": "user", "content": [
                    {"type": "text", "text": vision_prompt},
                    {"type": "image_url", "image_url": {"url": image_data}},
                ]}],
                max_tokens=600, temperature=0.1,
            ),
        )
        raw = re.sub(r"```json|```", "", vresp.choices[0].message.content.strip()).strip()
        vision_json = json.loads(raw)
    except Exception as exc:
        logger.error("[VISION] Stage 1: %s", exc)
        vision_json = {"description": "Image provided", "scene_type": "unknown"}

    memories  = await _get_active_memories(user_id, 15)
    mem_block = _format_memories(memories)
    combined  = (
        f"{mem_block}\n\n"
        f"## Image Analysis\n{json.dumps(vision_json, indent=2)}\n\n"
        f"## User Question\n{user_message}\n\nAnswer thoroughly."
    )
    answer = await _llm_text(
        [{"role": "user", "content": combined}],
        model_id="censored", max_tokens=1024, temperature=0.7,
    )
    return answer, json.dumps(vision_json)


# ══════════════════════════════════════════════════════════════════════════════
# §23  WEBSITE BUILDER
# ══════════════════════════════════════════════════════════════════════════════

_STYLE_HINTS = {
    "modern":        "Clean, bold typography, whitespace, subtle shadows, CSS Grid",
    "retro":         "80s/90s aesthetic, pixel fonts, neon on dark, CRT scanlines",
    "minimal":       "Extreme whitespace, single accent color, fine typography",
    "glassmorphism": "Frosted glass, backdrop-filter blur, translucency",
    "brutalist":     "Raw feel, bold borders, high contrast, asymmetric layout",
    "dark":          "Dark background, glowing accents, modern dark UI patterns",
    "futuristic":    "Sci-fi aesthetic, neon glows, animated particles, holographic effects",
}

_WEB_SYSTEM = """You are an elite frontend developer.
Generate a COMPLETE, SELF-CONTAINED single-file HTML website.
All CSS in <style>. All JS in <script>. CDN from Google Fonts / cdnjs.cloudflare.com only.
Fully functional, visually stunning, fully responsive. No Lorem Ipsum.
Return ONLY raw HTML — no markdown fences, no explanation."""

async def _build_website(description: str, title: str, style: str, model_id: str) -> str:
    hint   = _STYLE_HINTS.get(style, _STYLE_HINTS["modern"])
    prompt = f"Build a complete website.\nTitle: {title}\nDescription: {description}\nStyle: {style} — {hint}\nReturn ONLY raw HTML."
    result = await _llm_text(
        [{"role": "system", "content": _WEB_SYSTEM}, {"role": "user", "content": prompt}],
        model_id=model_id, max_tokens=4096, temperature=0.8,
    )
    html = re.sub(r"^```html\s*", "", result.strip(), flags=re.IGNORECASE)
    html = re.sub(r"```\s*$", "", html.strip())
    return html.strip()


# ══════════════════════════════════════════════════════════════════════════════
# §24  PYDANTIC MODELS
# ══════════════════════════════════════════════════════════════════════════════

class UserSignup(BaseModel):
    email:     str = Field(..., pattern=r"^[^@]+@[^@]+\.[^@]+$")
    password:  str = Field(..., min_length=8)
    full_name: Optional[str] = None

class UserLogin(BaseModel):
    email:    str
    password: str

class RefreshRequest(BaseModel):
    refresh_token: str

class ChangePasswordReq(BaseModel):
    current_password: str = Field(..., min_length=1)
    new_password:     str = Field(..., min_length=8)

class ProfileUpdate(BaseModel):
    full_name:      Optional[str]  = None
    memory_enabled: Optional[bool] = None
    timezone:       Optional[str]  = None

class ChatReq(BaseModel):
    message:         str  = Field(..., min_length=1, max_length=32000)
    model_id:        str  = Field("censored")
    use_rag:         bool = True
    image_url:       Optional[str] = None
    session_id:      Optional[str] = None
    force_thinking:  Optional[bool] = None

class RegenerateReq(BaseModel):
    message_id: str
    model_id:   str = "censored"

class EditMessageReq(BaseModel):
    message_id:  str
    new_content: str = Field(..., min_length=1, max_length=32000)

class BranchReq(BaseModel):
    session_id:         str
    branch_from_msg_id: str

class FeedbackReq(BaseModel):
    message_id: str
    rating:     int = Field(..., ge=-1, le=1)
    feedback:   str = ""

class MemoryManualCreate(BaseModel):
    key:   str = Field(..., min_length=1, max_length=40)
    value: str = Field(..., min_length=1, max_length=300)

class ConnectorCreate(BaseModel):
    name:           str
    connector_type: str = Field(..., pattern="^(slack|google_sheets|excel|power_bi|http)$")
    creds:          Dict[str, Any]

class AgentJobCreate(BaseModel):
    name:            str = Field(..., min_length=1, max_length=100)
    description:     str = ""
    job_type:        str = Field("cron", pattern="^(cron|manual|webhook)$")
    trigger_json:    str = "{}"
    prompt_template: str = Field(..., min_length=5)
    tools_json:      str = '["code_exec","shell","file_ops","rag_search","db_query"]'
    model_id:        str = "censored"
    max_retries:     int = Field(2, ge=0, le=5)
    timeout_seconds: int = Field(120, ge=10, le=600)

class AgentJobUpdate(BaseModel):
    name:            Optional[str]  = None
    prompt_template: Optional[str]  = None
    trigger_json:    Optional[str]  = None
    tools_json:      Optional[str]  = None
    enabled:         Optional[bool] = None
    model_id:        Optional[str]  = None

class AgentRunReq(BaseModel):
    prompt:          str  = Field(..., min_length=5, max_length=16000)
    tools:           List[str] = []  # empty = all tools allowed
    model_id:        str  = "censored"
    session_id:      Optional[str] = None
    timeout_seconds: int = Field(90, ge=10, le=300)

class ApiKeyCreate(BaseModel):
    name:           str = Field(..., min_length=1, max_length=60)
    scopes_json:    str = '["chat","rag"]'
    rate_limit_rpm: int = Field(60, ge=1, le=1000)
    expires_days:   Optional[int] = None

class WebsiteCreate(BaseModel):
    title:       str = Field(..., min_length=1, max_length=100)
    description: str = Field(..., min_length=5)
    style:       str = Field("modern")
    model_id:    str = "censored"

class CodeRunReq(BaseModel):
    code:     str = Field(..., min_length=1)
    language: str = Field("python")
    cwd:      Optional[str] = None

class ShellReq(BaseModel):
    command: str = Field(..., min_length=1)
    timeout: int = Field(30, ge=1, le=300)
    cwd:     Optional[str] = None

class FileOpsReq(BaseModel):
    op:      str
    path:    str
    content: Optional[str] = None

class VoiceTokenReq(BaseModel):
    room_name:   str  = "jazz-voice"
    identity:    Optional[str] = None
    ttl_seconds: int  = Field(3600, ge=60, le=86400)

class VoiceChatReq(BaseModel):
    transcript:   str
    session_id:   Optional[str] = None
    model_id:     str  = "censored"
    return_audio: bool = False
    voice:        str  = TTS_VOICE

class NotifCreate(BaseModel):
    user_id:      Optional[str] = None
    title:        str
    message:      str
    type:         str = "info"
    link:         Optional[str] = None
    is_broadcast: bool = False

class SubscriptionUpdate(BaseModel):
    user_id: str
    tier:    str = Field(..., pattern="^(free|pro|enterprise)$")

class CouponCreate(BaseModel):
    code:          str
    grants_tier:   str = Field(..., pattern="^(free|pro|enterprise)$")
    duration_days: int = 30
    max_uses:      int = 1
    expires_days:  Optional[int] = None


# ══════════════════════════════════════════════════════════════════════════════
# §25  APP FACTORY + LIFESPAN
# ══════════════════════════════════════════════════════════════════════════════

@asynccontextmanager
async def _lifespan(application: FastAPI) -> AsyncGenerator:
    logger.info("╔══ JAZZ AI v11 starting… ══╗")
    await _init_db_connection()
    await _bootstrap_schema()
    await _bootstrap_admin()
    _init_chroma()
    _start_scheduler()
    await _reload_all_jobs()
    logger.info("╚══ Ready ✅ ══╝")
    yield
    if _scheduler and _scheduler.running:
        _scheduler.shutdown(wait=False)
    if _db:
        await _db.close()
    logger.info("JAZZ AI v11 shutdown complete.")


app = FastAPI(
    title="JAZZ AI v11",
    version="11.0.0",
    description="Production-grade AI SaaS platform with full system access",
    lifespan=_lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_static_dir = Path("frontend")
if _static_dir.exists():
    app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")

_sites_dir = SITES_DIR
app.mount("/preview", StaticFiles(directory=str(_sites_dir)), name="sites")

@app.middleware("http")
async def _request_logger(request: Request, call_next):
    t0   = time.time()
    resp = await call_next(request)
    ms   = int((time.time() - t0) * 1000)
    if request.url.path not in ("/health", "/favicon.ico"):
        logger.info("%s %s → %d (%dms)", request.method, request.url.path, resp.status_code, ms)
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# §26  AUTH ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/auth/signup")
async def signup(req: UserSignup) -> Dict:
    if await db_fetchone("SELECT id FROM users WHERE email=?", (req.email,)):
        raise HTTPException(400, "Email already registered")
    uid = _new_id(); now = _utcnow()
    ph  = _pwd.hash(_truncate_password(req.password))
    await db_execute(
        "INSERT INTO users (id,email,password_hash,full_name,role,subscription,created_at,updated_at)"
        " VALUES (?,?,?,?,?,?,?,?)",
        (uid, req.email, ph, req.full_name or "", "client", "free", now, now),
    )
    access  = _create_jwt({"sub": uid, "email": req.email, "role": "client", "subscription": "free"}, ACCESS_TOKEN_EXPIRE_MINUTES)
    refresh = await _create_refresh_token(uid)
    return {"access_token": access, "refresh_token": refresh, "token_type": "bearer",
            "user": {"id": uid, "email": req.email, "role": "client", "subscription": "free"}}

@app.post("/auth/login")
async def login(req: UserLogin) -> Dict:
    user = await db_fetchone("SELECT * FROM users WHERE email=? AND is_active=1", (req.email,))
    if not user or not _pwd.verify(_truncate_password(req.password), user["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    await db_execute("UPDATE users SET last_login_at=? WHERE id=?", (_utcnow(), user["id"]))
    payload = {"sub": user["id"], "email": user["email"], "role": user["role"], "subscription": user["subscription"]}
    access  = _create_jwt(payload, ACCESS_TOKEN_EXPIRE_MINUTES)
    refresh = await _create_refresh_token(user["id"])
    return {"access_token": access, "refresh_token": refresh, "token_type": "bearer",
            "user": {k: user[k] for k in ("id","email","full_name","role","subscription","memory_enabled")}}

@app.post("/auth/refresh")
async def refresh_token(req: RefreshRequest) -> Dict:
    user_id = await _validate_refresh_token(req.refresh_token)
    if not user_id:
        raise HTTPException(401, "Invalid or expired refresh token")
    user = await db_fetchone("SELECT * FROM users WHERE id=? AND is_active=1", (user_id,))
    if not user:
        raise HTTPException(401, "User not found")
    await _revoke_refresh_token(req.refresh_token)
    payload = {"sub": user["id"], "email": user["email"], "role": user["role"], "subscription": user["subscription"]}
    access  = _create_jwt(payload, ACCESS_TOKEN_EXPIRE_MINUTES)
    new_refresh = await _create_refresh_token(user["id"])
    return {"access_token": access, "refresh_token": new_refresh, "token_type": "bearer"}

@app.post("/auth/logout")
async def logout(req: RefreshRequest) -> Dict:
    await _revoke_refresh_token(req.refresh_token)
    return {"success": True}

@app.get("/auth/me")
async def me(user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone(
        "SELECT id,email,full_name,role,subscription,memory_enabled,timezone,created_at FROM users WHERE id=?",
        (user["sub"],),
    )
    return row or {}

@app.put("/auth/profile")
async def update_profile(req: ProfileUpdate, user: Dict = Depends(get_current_user)) -> Dict:
    updates, vals = [], []
    if req.full_name      is not None: updates.append("full_name=?"); vals.append(req.full_name)
    if req.memory_enabled is not None: updates.append("memory_enabled=?"); vals.append(int(req.memory_enabled))
    if req.timezone       is not None: updates.append("timezone=?"); vals.append(req.timezone)
    if updates:
        updates.append("updated_at=?"); vals.append(_utcnow()); vals.append(user["sub"])
        await db_execute(f"UPDATE users SET {','.join(updates)} WHERE id=?", tuple(vals))
    return {"success": True}

@app.post("/auth/change-password")
async def change_password(req: ChangePasswordReq, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT password_hash FROM users WHERE id=?", (user["sub"],))
    if not row or not _pwd.verify(_truncate_password(req.current_password), row["password_hash"]):
        raise HTTPException(400, "Current password incorrect")
    await db_execute(
        "UPDATE users SET password_hash=?, updated_at=? WHERE id=?",
        (_pwd.hash(_truncate_password(req.new_password)), _utcnow(), user["sub"]),
    )
    return {"success": True}


# ══════════════════════════════════════════════════════════════════════════════
# §27  CHAT ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/chat")
async def chat_endpoint(
    req: ChatReq, background: BackgroundTasks,
    user: Dict = Depends(_require_scope("chat")),
) -> Dict:
    await _check_rate(user, "messages_per_day")
    if req.image_url:
        answer, vision_meta = await _analyze_image(req.image_url, req.message, user["sub"])
        return {"response": answer, "vision_meta": vision_meta, "agent": "vision_agent"}
    return await chat_send_message(
        user["sub"], req.message, req.session_id,
        req.model_id, req.use_rag, background, req.force_thinking,
    )

@app.get("/chat/stream")
async def chat_stream_get(
    message: str, model_id: str = "censored", use_rag: bool = True,
    session_id: Optional[str] = None, force_thinking: Optional[bool] = None,
    user: Dict = Depends(get_current_user),
) -> StreamingResponse:
    await _check_rate(user, "messages_per_day")
    async def _gen():
        async for chunk in chat_stream_message(
            user["sub"], message, session_id, model_id, use_rag, force_thinking
        ):
            yield chunk
    return StreamingResponse(
        _gen(), media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )

@app.post("/chat/stream")
async def chat_stream_post(req: ChatReq, user: Dict = Depends(get_current_user)) -> StreamingResponse:
    await _check_rate(user, "messages_per_day")
    async def _gen():
        async for chunk in chat_stream_message(
            user["sub"], req.message, req.session_id, req.model_id, req.use_rag, req.force_thinking
        ):
            yield chunk
    return StreamingResponse(
        _gen(), media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )

@app.post("/chat/regenerate")
async def regenerate(req: RegenerateReq, user: Dict = Depends(get_current_user)) -> Dict:
    await _check_rate(user, "messages_per_day")
    row = await db_fetchone(
        "SELECT session_id, content FROM chat_history WHERE id=? AND user_id=? AND role='user'",
        (req.message_id, user["sub"]),
    )
    if not row:
        raise HTTPException(404, "Message not found")
    return await chat_send_message(user["sub"], row["content"], row["session_id"], req.model_id, use_rag=True)

@app.post("/chat/edit")
async def edit_message(req: EditMessageReq, user: Dict = Depends(get_current_user)) -> Dict:
    await _check_rate(user, "messages_per_day")
    row = await db_fetchone(
        "SELECT session_id, content, created_at FROM chat_history WHERE id=? AND user_id=? AND role='user'",
        (req.message_id, user["sub"]),
    )
    if not row:
        raise HTTPException(404, "Message not found")
    await db_execute(
        "INSERT INTO message_edits (id,message_id,old_content,new_content,edited_at) VALUES (?,?,?,?,?)",
        (_new_id(), req.message_id, row["content"], req.new_content, _utcnow()),
    )
    await db_execute("UPDATE chat_history SET content=?, edit_count=edit_count+1 WHERE id=?", (req.new_content, req.message_id))
    await db_execute(
        "UPDATE chat_history SET is_hidden=1 WHERE session_id=? AND created_at>? AND id!=?",
        (row["session_id"], row["created_at"], req.message_id),
    )
    return await chat_send_message(user["sub"], req.new_content, row["session_id"], use_rag=True)

@app.post("/chat/branch")
async def branch_conversation(req: BranchReq, user: Dict = Depends(get_current_user)) -> Dict:
    parent = await db_fetchone("SELECT title FROM chat_sessions WHERE id=? AND user_id=?", (req.session_id, user["sub"]))
    if not parent:
        raise HTTPException(404, "Session not found")
    new_sid = _new_id(); now = _utcnow()
    await db_execute(
        "INSERT INTO chat_sessions (id,user_id,title,parent_session_id,branch_from_msg_id,created_at,updated_at)"
        " VALUES (?,?,?,?,?,?,?)",
        (new_sid, user["sub"], f"{parent['title']} (branch)", req.session_id, req.branch_from_msg_id, now, now),
    )
    return {"new_session_id": new_sid}

@app.post("/chat/feedback")
async def message_feedback(req: FeedbackReq, user: Dict = Depends(get_current_user)) -> Dict:
    await db_execute(
        "INSERT OR REPLACE INTO message_feedback (id,message_id,user_id,rating,feedback,created_at) VALUES (?,?,?,?,?,?)",
        (_new_id(), req.message_id, user["sub"], req.rating, req.feedback, _utcnow()),
    )
    return {"success": True}

@app.get("/sessions")
async def list_sessions(limit: int = 60, offset: int = 0, user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,title,model_id,turn_count,last_message_at,is_pinned,is_archived,created_at"
        " FROM chat_sessions WHERE user_id=? AND is_archived=0"
        " ORDER BY is_pinned DESC, last_message_at DESC LIMIT ? OFFSET ?",
        (user["sub"], limit, offset),
    )
    return {"sessions": rows}

@app.get("/sessions/{session_id}/messages")
async def get_session_messages(session_id: str, limit: int = 60, offset: int = 0, user: Dict = Depends(get_current_user)) -> Dict:
    msgs = await db_fetchall(
        "SELECT id,role,content,model_used,agent_used,tokens_input,tokens_output,"
        " latency_ms,edit_count,thinking_content,mode,created_at"
        " FROM chat_history WHERE session_id=? AND user_id=? AND is_hidden=0"
        " ORDER BY created_at ASC LIMIT ? OFFSET ?",
        (session_id, user["sub"], limit, offset),
    )
    summaries = await db_fetchall(
        "SELECT summary_text, covers_from_at, covers_to_at FROM summaries WHERE session_id=? ORDER BY created_at",
        (session_id,),
    )
    return {"messages": msgs, "summaries": summaries}

@app.delete("/sessions/{session_id}")
async def delete_session(session_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    await db_execute("DELETE FROM chat_sessions WHERE id=? AND user_id=?", (session_id, user["sub"]))
    return {"success": True}

@app.patch("/sessions/{session_id}/pin")
async def pin_session(session_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT is_pinned FROM chat_sessions WHERE id=? AND user_id=?", (session_id, user["sub"]))
    if not row:
        raise HTTPException(404, "Session not found")
    new_pin = 0 if row["is_pinned"] else 1
    await db_execute("UPDATE chat_sessions SET is_pinned=?, updated_at=? WHERE id=?", (new_pin, _utcnow(), session_id))
    return {"is_pinned": bool(new_pin)}

@app.get("/sessions/{session_id}/export")
async def export_session(session_id: str, format: str = "json", user: Dict = Depends(get_current_user)) -> Any:
    msgs = await db_fetchall(
        "SELECT role,content,created_at,agent_used FROM chat_history"
        " WHERE session_id=? AND user_id=? AND is_hidden=0 ORDER BY created_at",
        (session_id, user["sub"]),
    )
    if not msgs:
        raise HTTPException(404, "No messages found")
    if format == "json":
        return JSONResponse({"messages": msgs})
    elif format == "markdown":
        lines = ["# Chat Export\n"]
        for m in msgs:
            lines.append(f"**{m['role'].capitalize()}** ({m['created_at']}):\n{m['content']}\n")
        return StreamingResponse(io.StringIO("\n".join(lines)), media_type="text/markdown",
                                 headers={"Content-Disposition": f'attachment; filename="chat_{session_id[:8]}.md"'})
    raise HTTPException(400, "format must be json|markdown")


# ══════════════════════════════════════════════════════════════════════════════
# §28  DOCUMENTS
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/documents/upload")
async def upload_document(background: BackgroundTasks, file: UploadFile = File(...), user: Dict = Depends(_require_scope("rag"))) -> Dict:
    await _check_rate(user, "documents")
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, "File too large")
    doc_id   = _new_id()
    ext      = Path(file.filename).suffix
    fname    = f"{doc_id}{ext}"
    fpath    = UPLOADS_DIR / fname
    fpath.write_bytes(content)
    now = _utcnow()
    await db_execute(
        "INSERT INTO documents (id,user_id,filename,original_name,file_type,file_size_bytes,collection_name,uploaded_at)"
        " VALUES (?,?,?,?,?,?,?,?)",
        (doc_id, user["sub"], fname, file.filename, file.content_type or ext, len(content), "documents", now),
    )
    background.add_task(_ingest_document, doc_id, str(fpath), file.content_type or "", user["sub"])
    return {"doc_id": doc_id, "filename": file.filename, "status": "indexing"}

@app.get("/documents")
async def list_documents(user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,original_name,file_type,file_size_bytes,chunk_count,is_indexed,index_error,uploaded_at,indexed_at"
        " FROM documents WHERE user_id=? ORDER BY uploaded_at DESC",
        (user["sub"],),
    )
    return {"documents": rows}

@app.delete("/documents/{doc_id}")
async def delete_document(doc_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT filename FROM documents WHERE id=? AND user_id=?", (doc_id, user["sub"]))
    if not row:
        raise HTTPException(404, "Document not found")
    if _docs_collection:
        try: _docs_collection.delete(where={"doc_id": doc_id})
        except Exception: pass
    fpath = UPLOADS_DIR / row["filename"]
    if fpath.exists(): fpath.unlink()
    await db_execute("DELETE FROM documents WHERE id=?", (doc_id,))
    await db_execute("DELETE FROM embeddings_meta WHERE document_id=?", (doc_id,))
    return {"success": True}


# ══════════════════════════════════════════════════════════════════════════════
# §29  MEMORY
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/memories")
async def get_memories(user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,key,value,source,confidence,reinforcement_count,last_reinforced,created_at"
        " FROM memories WHERE user_id=? AND is_active=1"
        " ORDER BY confidence DESC, last_reinforced DESC",
        (user["sub"],),
    )
    return {"memories": rows}

@app.post("/memories")
async def add_memory(req: MemoryManualCreate, user: Dict = Depends(get_current_user)) -> Dict:
    await _upsert_memory(user["sub"], req.key, req.value, "manual", 1.0)
    return {"success": True}

@app.delete("/memories/{memory_id}")
async def delete_memory(memory_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    await db_execute("UPDATE memories SET is_active=0 WHERE id=? AND user_id=?", (memory_id, user["sub"]))
    return {"success": True}

@app.delete("/memories")
async def clear_memories(user: Dict = Depends(get_current_user)) -> Dict:
    await db_execute("UPDATE memories SET is_active=0 WHERE user_id=?", (user["sub"],))
    return {"success": True}


# ══════════════════════════════════════════════════════════════════════════════
# §30  API KEYS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api-keys")
async def list_api_keys(user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,name,key_prefix,scopes_json,rate_limit_rpm,is_active,last_used_at,usage_count,expires_at,created_at"
        " FROM api_keys WHERE user_id=? ORDER BY created_at DESC",
        (user["sub"],),
    )
    return {"api_keys": rows}

@app.post("/api-keys")
async def create_api_key(req: ApiKeyCreate, user: Dict = Depends(get_current_user)) -> Dict:
    raw    = f"jz_{secrets.token_urlsafe(36)}"
    h      = _hash_token(raw)
    prefix = raw[:10]
    exp    = None
    if req.expires_days:
        exp = (datetime.now(timezone.utc) + timedelta(days=req.expires_days)).isoformat()
    kid = _new_id()
    await db_execute(
        "INSERT INTO api_keys (id,user_id,name,key_hash,key_prefix,scopes_json,rate_limit_rpm,expires_at,created_at)"
        " VALUES (?,?,?,?,?,?,?,?,?)",
        (kid, user["sub"], req.name, h, prefix, req.scopes_json, req.rate_limit_rpm, exp, _utcnow()),
    )
    return {"id": kid, "key": raw, "prefix": prefix, "name": req.name}

@app.delete("/api-keys/{key_id}")
async def revoke_api_key(key_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    await db_execute("UPDATE api_keys SET is_active=0 WHERE id=? AND user_id=?", (key_id, user["sub"]))
    return {"success": True}


# ══════════════════════════════════════════════════════════════════════════════
# §31  AGENT JOBS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/agent-jobs")
async def list_agent_jobs(user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,name,description,job_type,status,enabled,total_runs,success_runs,"
        "failed_runs,last_run_at,last_run_status,next_run_at,created_at"
        " FROM agent_jobs WHERE user_id=? ORDER BY created_at DESC",
        (user["sub"],),
    )
    return {"jobs": rows}

@app.post("/agent-jobs")
async def create_agent_job(req: AgentJobCreate, user: Dict = Depends(get_current_user)) -> Dict:
    jid = _new_id(); now = _utcnow()
    await db_execute(
        "INSERT INTO agent_jobs (id,user_id,name,description,job_type,trigger_json,"
        "prompt_template,tools_json,model_id,max_retries,timeout_seconds,created_at,updated_at)"
        " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (jid, user["sub"], req.name, req.description, req.job_type, req.trigger_json,
         req.prompt_template, req.tools_json, req.model_id, req.max_retries, req.timeout_seconds, now, now),
    )
    job_row = await db_fetchone("SELECT * FROM agent_jobs WHERE id=?", (jid,))
    _schedule_job(job_row)
    return {"id": jid, "status": "created"}

@app.put("/agent-jobs/{job_id}")
async def update_agent_job(job_id: str, req: AgentJobUpdate, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT id FROM agent_jobs WHERE id=? AND user_id=?", (job_id, user["sub"]))
    if not row:
        raise HTTPException(404, "Job not found")
    updates, vals = [], []
    for field, col in [("name","name"),("prompt_template","prompt_template"),("trigger_json","trigger_json"),("tools_json","tools_json"),("model_id","model_id")]:
        val = getattr(req, field)
        if val is not None: updates.append(f"{col}=?"); vals.append(val)
    if req.enabled is not None: updates.append("enabled=?"); vals.append(int(req.enabled))
    if updates:
        updates.append("updated_at=?"); vals.append(_utcnow()); vals.append(job_id)
        await db_execute(f"UPDATE agent_jobs SET {','.join(updates)} WHERE id=?", tuple(vals))
    job_row = await db_fetchone("SELECT * FROM agent_jobs WHERE id=?", (job_id,))
    _schedule_job(job_row)
    return {"success": True}

@app.post("/agent-jobs/{job_id}/run")
async def run_job_now(job_id: str, background: BackgroundTasks, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT id FROM agent_jobs WHERE id=? AND user_id=?", (job_id, user["sub"]))
    if not row: raise HTTPException(404, "Job not found")
    background.add_task(_run_job, job_id)
    return {"status": "queued"}

@app.delete("/agent-jobs/{job_id}")
async def delete_agent_job(job_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT id FROM agent_jobs WHERE id=? AND user_id=?", (job_id, user["sub"]))
    if not row: raise HTTPException(404, "Job not found")
    if _scheduler:
        try: _scheduler.remove_job(job_id)
        except Exception: pass
    await db_execute("DELETE FROM agent_jobs WHERE id=?", (job_id,))
    return {"success": True}

@app.get("/agent-jobs/{job_id}/logs")
async def job_logs(job_id: str, limit: int = 20, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT id FROM agent_jobs WHERE id=? AND user_id=?", (job_id, user["sub"]))
    if not row: raise HTTPException(404, "Job not found")
    logs = await db_fetchall(
        "SELECT id,run_number,status,trigger_type,tokens_used,latency_ms,"
        "error_message,started_at,finished_at,result_text FROM agent_job_logs"
        " WHERE job_id=? ORDER BY started_at DESC LIMIT ?",
        (job_id, limit),
    )
    return {"logs": logs}


# ══════════════════════════════════════════════════════════════════════════════
# §32  AGENT INLINE
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/agent/run")
async def agent_run(req: AgentRunReq, background: BackgroundTasks, user: Dict = Depends(get_current_user)) -> Dict:
    await _check_rate(user, "messages_per_day")
    t0 = time.time()
    answer, tool_log = await _run_agent_loop(
        user_id=user["sub"], prompt=req.prompt,
        allowed_tools=req.tools, model_id=req.model_id, timeout_s=req.timeout_seconds,
    )
    elapsed = int((time.time() - t0) * 1000)
    return {"answer": answer, "tool_calls": tool_log, "steps": len(tool_log), "latency_ms": elapsed}

@app.post("/agent/run/stream")
async def agent_run_stream(req: AgentRunReq, user: Dict = Depends(get_current_user)) -> StreamingResponse:
    await _check_rate(user, "messages_per_day")

    async def _gen():
        yield f"data: {json.dumps({'type':'start','prompt':req.prompt[:100]})}\n\n"
        messages: List[Dict] = [
            {"role": "system", "content": _AGENT_SYSTEM},
            {"role": "user", "content": req.prompt},
        ]
        tool_log: List[Dict] = []
        deadline = time.time() + req.timeout_seconds

        for step in range(_MAX_AGENT_STEPS):
            if time.time() > deadline:
                yield f"data: {json.dumps({'type':'timeout'})}\n\n"
                break
            try:
                raw       = await _llm_text(messages, req.model_id, 800, 0.2)
                clean     = re.sub(r"```json|```", "", raw.strip()).strip()
                step_json = json.loads(clean)
            except Exception as exc:
                yield f"data: {json.dumps({'type':'parse_error','error':str(exc)})}\n\n"
                break

            thought = step_json.get("thought", "")
            action  = step_json.get("action", "FINISH")
            args    = step_json.get("args", {})
            messages.append({"role": "assistant", "content": raw})
            yield f"data: {json.dumps({'type':'thought','step':step,'thought':thought,'action':action})}\n\n"

            if action == "FINISH":
                final = step_json.get("final_answer", "Done.")
                yield f"data: {json.dumps({'type':'done','answer':final,'steps':len(tool_log)})}\n\n"
                return

            t_start = time.time()
            result  = await _dispatch_tool(action, args, user["sub"])
            obs     = json.dumps({"success": result.success, "output": str(result.output)[:2000] if result.output else None, "error": result.error})
            tool_log.append({"step": step, "tool": action, "success": result.success})
            yield f"data: {json.dumps({'type':'tool_result','tool':action,'success':result.success,'latency_ms':int((time.time()-t_start)*1000),'output':str(result.output)[:500] if result.output else None})}\n\n"
            messages.append({"role": "user", "content": f"[Observation]: {obs}"})

        yield f"data: {json.dumps({'type':'done','answer':'Agent loop ended.','steps':len(tool_log)})}\n\n"

    return StreamingResponse(_gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ══════════════════════════════════════════════════════════════════════════════
# §33  COMPUTER / SYSTEM ACCESS ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/system/shell")
async def run_shell_cmd(req: ShellReq, user: Dict = Depends(get_current_user)) -> Dict:
    """Execute any shell command on the host system."""
    result = await asyncio.get_event_loop().run_in_executor(
        _executor, lambda: _run_shell(req.command, req.timeout, req.cwd)
    )
    return result

@app.post("/system/code")
async def run_code(req: CodeRunReq, user: Dict = Depends(get_current_user)) -> Dict:
    """Run code in any supported language."""
    await _check_rate(user, "code_runs_per_day")
    result = await asyncio.get_event_loop().run_in_executor(
        _executor, lambda: _run_code_sync(req.code, req.language, req.cwd)
    )
    return result

@app.post("/system/files")
async def file_operations(req: FileOpsReq, user: Dict = Depends(get_current_user)) -> Dict:
    """File system operations: list, read, write, delete."""
    result_tool = await _tool_file_ops({"op": req.op, "path": req.path, "content": req.content})
    return {"success": result_tool.success, "output": result_tool.output, "error": result_tool.error}

@app.get("/system/info")
async def system_info_ep(user: Dict = Depends(get_current_user)) -> Dict:
    """Get system/OS information."""
    return await asyncio.get_event_loop().run_in_executor(_executor, _get_system_info)

@app.post("/system/open")
async def open_app_ep(app_name: str, user: Dict = Depends(get_current_user)) -> Dict:
    """Open an application or URL on the host."""
    result = await asyncio.get_event_loop().run_in_executor(_executor, lambda: _open_application(app_name))
    return result

# Legacy route compatibility
@app.post("/code/run")
async def run_code_legacy(req: CodeRunReq, user: Dict = Depends(get_current_user)) -> Dict:
    await _check_rate(user, "code_runs_per_day")
    result = await asyncio.get_event_loop().run_in_executor(
        _executor, lambda: _run_code_sync(req.code, req.language, req.cwd)
    )
    return result


# ══════════════════════════════════════════════════════════════════════════════
# §34  CONNECTORS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/connectors")
async def list_connectors(user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,name,connector_type,is_active,last_tested_at,last_test_ok,test_error,created_at"
        " FROM connectors WHERE user_id=? ORDER BY created_at DESC",
        (user["sub"],),
    )
    return {"connectors": rows}

@app.post("/connectors")
async def create_connector(req: ConnectorCreate, user: Dict = Depends(get_current_user)) -> Dict:
    cid = _new_id(); now = _utcnow()
    enc = _encrypt(req.creds)
    try:
        await db_execute(
            "INSERT INTO connectors (id,user_id,name,connector_type,encrypted_creds,created_at,updated_at)"
            " VALUES (?,?,?,?,?,?,?)",
            (cid, user["sub"], req.name, req.connector_type, enc, now, now),
        )
    except Exception:
        raise HTTPException(400, f"Connector named '{req.name}' already exists")
    return {"id": cid, "name": req.name, "connector_type": req.connector_type}

@app.post("/connectors/{connector_id}/test")
async def test_connector(connector_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT * FROM connectors WHERE id=? AND user_id=?", (connector_id, user["sub"]))
    if not row: raise HTTPException(404, "Connector not found")
    try:
        creds = _decrypt(row["encrypted_creds"])
        test_actions = {
            "slack":         ("list_channels", {}),
            "google_sheets": ("get_metadata", {}),
            "power_bi":      ("list_datasets", {}),
            "http":          ("GET", {"path": "/", "method": "GET"}),
        }
        action, params = test_actions.get(row["connector_type"], ("GET", {}))
        await _connector_dispatch(row["connector_type"], action, creds, params)
        await db_execute("UPDATE connectors SET last_tested_at=?, last_test_ok=1, test_error=NULL WHERE id=?", (_utcnow(), connector_id))
        return {"success": True}
    except Exception as exc:
        await db_execute("UPDATE connectors SET last_tested_at=?, last_test_ok=0, test_error=? WHERE id=?", (_utcnow(), str(exc)[:300], connector_id))
        return {"success": False, "error": str(exc)}

@app.delete("/connectors/{connector_id}")
async def delete_connector(connector_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    await db_execute("DELETE FROM connectors WHERE id=? AND user_id=?", (connector_id, user["sub"]))
    return {"success": True}


# ══════════════════════════════════════════════════════════════════════════════
# §35  VOICE / LIVEKIT ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/voice/token")
async def voice_token(req: VoiceTokenReq, user: Dict = Depends(get_current_user)) -> Dict:
    identity = req.identity or user["sub"]
    name     = user.get("email", identity)
    try:
        token = _lk_token(identity, name, req.room_name, req.ttl_seconds)
        return {
            "token": token, "room": req.room_name,
            "url": LIVEKIT_URL, "identity": identity,
            "expires_in": req.ttl_seconds,
        }
    except Exception as exc:
        raise HTTPException(500, f"Token generation failed: {exc}")

@app.post("/voice/transcribe")
async def transcribe_audio(file: UploadFile = File(...), user: Dict = Depends(get_current_user)) -> Dict:
    await _check_rate(user, "messages_per_day")
    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(413, "Audio file too large (max 25 MB)")
    transcript = await _stt_transcribe(content, file.content_type or "audio/webm")
    return {"transcript": transcript, "chars": len(transcript)}

@app.post("/voice/chat")
async def voice_chat(req: VoiceChatReq, background: BackgroundTasks, user: Dict = Depends(get_current_user)) -> Dict:
    await _check_rate(user, "messages_per_day")
    result = await chat_send_message(
        user["sub"], req.transcript, req.session_id,
        req.model_id, use_rag=True, background=background,
    )
    response: Dict = {
        "session_id": result["session_id"], "message_id": result["message_id"],
        "response": result["response"], "agent": result["agent"], "latency_ms": result["latency_ms"],
    }
    if req.return_audio:
        try:
            audio_bytes = await _tts_synthesize(result["response"], req.voice)
            response["audio_base64"] = base64.b64encode(audio_bytes).decode()
            response["audio_format"] = "mp3"
        except Exception as exc:
            response["tts_error"] = str(exc)
    return response

@app.post("/voice/tts")
async def text_to_speech(text: str, voice: str = TTS_VOICE, user: Dict = Depends(get_current_user)) -> StreamingResponse:
    await _check_rate(user, "messages_per_day")
    if not text.strip():
        raise HTTPException(400, "text is required")
    audio_bytes = await _tts_synthesize(text[:4096], voice)
    return StreamingResponse(
        io.BytesIO(audio_bytes), media_type="audio/mpeg",
        headers={"Content-Disposition": "inline; filename=speech.mp3"},
    )

@app.post("/voice/stream-chat")
async def voice_stream_chat(req: VoiceChatReq, user: Dict = Depends(get_current_user)) -> StreamingResponse:
    await _check_rate(user, "messages_per_day")
    async def _gen():
        async for chunk in chat_stream_message(user["sub"], req.transcript, req.session_id, req.model_id, True):
            yield chunk
    return StreamingResponse(_gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ══════════════════════════════════════════════════════════════════════════════
# §36  WEBSOCKET
# ══════════════════════════════════════════════════════════════════════════════

@app.websocket("/ws/{user_id}")
async def websocket_endpoint(user_id: str, ws: WebSocket) -> None:
    await ws.accept()
    try:
        auth_msg = await asyncio.wait_for(ws.receive_json(), timeout=15.0)
    except asyncio.TimeoutError:
        await ws.close(code=4001, reason="Auth timeout")
        return

    token   = auth_msg.get("token", "")
    payload = _decode_jwt(token) if token else None
    if not payload or payload.get("sub") != user_id:
        await ws.send_json({"type": "error", "message": "Unauthorized"})
        await ws.close(code=4003, reason="Unauthorized")
        return

    ws_manager._connections[user_id].add(ws)
    await ws.send_json({"type": "connected", "user_id": user_id})

    try:
        while True:
            msg   = await ws.receive_json()
            mtype = msg.get("type", "")
            if mtype == "ping":
                await ws.send_json({"type": "pong", "ts": _utcnow()})
            elif mtype == "typing":
                await ws.send_json({"type": "typing_ack", "session_id": msg.get("session_id", "")})
            elif mtype == "presence":
                await ws.send_json({"type": "presence", "online": len(ws_manager.online_users())})
    except WebSocketDisconnect:
        pass
    except Exception as exc:
        logger.warning("[WS] %s error: %s", user_id[:8], exc)
    finally:
        ws_manager.disconnect(user_id, ws)


# ══════════════════════════════════════════════════════════════════════════════
# §37  WEBSITES
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/websites/build")
async def build_website(req: WebsiteCreate, user: Dict = Depends(get_current_user)) -> Dict:
    await _check_rate(user, "websites")
    html  = await _build_website(req.description, req.title, req.style, req.model_id)
    wid   = _new_id(); now = _utcnow()
    fname = f"{wid}.html"
    fpath = SITES_DIR / fname
    fpath.write_text(html, encoding="utf-8")
    await db_execute(
        "INSERT INTO websites (id,user_id,title,description,filename,style,prompt_used,"
        "html_size_bytes,created_at,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (wid, user["sub"], req.title, req.description, fname, req.style, req.description, len(html.encode()), now, now),
    )
    return {"id": wid, "title": req.title, "preview_url": f"/preview/{fname}"}

@app.get("/websites")
async def list_websites(user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,title,description,filename,style,html_size_bytes,view_count,created_at"
        " FROM websites WHERE user_id=? ORDER BY created_at DESC",
        (user["sub"],),
    )
    for r in rows: r["preview_url"] = f"/preview/{r['filename']}"
    return {"websites": rows}

@app.delete("/websites/{site_id}")
async def delete_website(site_id: str, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone("SELECT filename FROM websites WHERE id=? AND user_id=?", (site_id, user["sub"]))
    if not row: raise HTTPException(404, "Website not found")
    fpath = SITES_DIR / row["filename"]
    if fpath.exists(): fpath.unlink()
    await db_execute("DELETE FROM websites WHERE id=?", (site_id,))
    return {"success": True}


# ══════════════════════════════════════════════════════════════════════════════
# §38  NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/notifications")
async def get_notifications(user: Dict = Depends(get_current_user)) -> Dict:
    rows = await db_fetchall(
        "SELECT id,title,message,type,link,is_read,created_at FROM notifications"
        " WHERE (user_id=? OR is_broadcast=1) AND (expires_at IS NULL OR expires_at > ?)"
        " ORDER BY created_at DESC LIMIT 50",
        (user["sub"], _utcnow()),
    )
    return {"notifications": rows, "unread": sum(1 for r in rows if not r["is_read"])}

@app.post("/notifications/read-all")
async def mark_all_read(user: Dict = Depends(get_current_user)) -> Dict:
    await db_execute("UPDATE notifications SET is_read=1 WHERE user_id=?", (user["sub"],))
    return {"success": True}


# ══════════════════════════════════════════════════════════════════════════════
# §39  COUPONS
# ══════════════════════════════════════════════════════════════════════════════

@app.post("/coupons/redeem")
async def redeem_coupon(code: str, user: Dict = Depends(get_current_user)) -> Dict:
    row = await db_fetchone(
        "SELECT * FROM coupons WHERE code=? AND is_active=1 AND (expires_at IS NULL OR expires_at > ?)",
        (code, _utcnow()),
    )
    if not row: raise HTTPException(404, "Invalid or expired coupon code")
    if row["current_uses"] >= row["max_uses"]: raise HTTPException(400, "Coupon fully used")
    already = await db_fetchone("SELECT id FROM coupon_redemptions WHERE coupon_id=? AND user_id=?", (row["id"], user["sub"]))
    if already: raise HTTPException(400, "Already redeemed")
    exp = (datetime.now(timezone.utc) + timedelta(days=row["duration_days"])).isoformat()
    async with db_transaction():
        await db_execute("UPDATE users SET subscription=?, subscription_expires_at=?, updated_at=? WHERE id=?", (row["grants_tier"], exp, _utcnow(), user["sub"]))
        await db_execute("UPDATE coupons SET current_uses=current_uses+1 WHERE id=?", (row["id"],))
        await db_execute("INSERT INTO coupon_redemptions (id,coupon_id,user_id,redeemed_at) VALUES (?,?,?,?)", (_new_id(), row["id"], user["sub"], _utcnow()))
    return {"success": True, "new_tier": row["grants_tier"], "expires_at": exp}


# ══════════════════════════════════════════════════════════════════════════════
# §40  USAGE & SEARCH
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/usage")
async def usage_stats(user: Dict = Depends(get_current_user)) -> Dict:
    sub = user.get("subscription", "free")
    lim = SUBSCRIPTION_LIMITS[sub]
    return {
        "subscription": sub, "limits": lim,
        "usage": {
            "messages_today":  await rate_limiter.usage(user["sub"], "messages_per_day"),
            "code_runs_today": await rate_limiter.usage(user["sub"], "code_runs_per_day"),
            "documents":       await db_count("SELECT COUNT(*) FROM documents WHERE user_id=?", (user["sub"],)),
            "memories":        await db_count("SELECT COUNT(*) FROM memories WHERE user_id=? AND is_active=1", (user["sub"],)),
            "agent_jobs":      await db_count("SELECT COUNT(*) FROM agent_jobs WHERE user_id=?", (user["sub"],)),
            "connectors":      await db_count("SELECT COUNT(*) FROM connectors WHERE user_id=?", (user["sub"],)),
        },
    }

@app.get("/search")
async def search(q: str, scope: str = "messages", limit: int = 20, user: Dict = Depends(get_current_user)) -> Dict:
    if len(q.strip()) < 2: raise HTTPException(400, "Query must be at least 2 characters")
    pattern = f"%{q}%"
    results: Dict[str, List] = {}
    if scope in ("messages", "all"):
        rows = await db_fetchall(
            "SELECT ch.id, ch.role, ch.content, ch.created_at, cs.id as session_id, cs.title as session_title"
            " FROM chat_history ch JOIN chat_sessions cs ON ch.session_id=cs.id"
            " WHERE ch.user_id=? AND ch.is_hidden=0 AND ch.content LIKE ?"
            " ORDER BY ch.created_at DESC LIMIT ?",
            (user["sub"], pattern, limit),
        )
        results["messages"] = rows
    if scope in ("memories", "all"):
        results["memories"] = await db_fetchall(
            "SELECT id, key, value, source, confidence FROM memories"
            " WHERE user_id=? AND is_active=1 AND (key LIKE ? OR value LIKE ?) LIMIT ?",
            (user["sub"], pattern, pattern, limit),
        )
    return {"query": q, "scope": scope, "results": results}


# ══════════════════════════════════════════════════════════════════════════════
# §41  ADMIN
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/users")
async def admin_list_users(limit: int = 100, offset: int = 0, admin: Dict = Depends(require_admin)) -> Dict:
    rows  = await db_fetchall("SELECT id,email,full_name,role,subscription,is_active,last_login_at,created_at FROM users ORDER BY created_at DESC LIMIT ? OFFSET ?", (limit, offset))
    total = await db_count("SELECT COUNT(*) FROM users")
    return {"users": rows, "total": total}

@app.put("/admin/subscription")
async def admin_update_subscription(req: SubscriptionUpdate, admin: Dict = Depends(require_admin)) -> Dict:
    await db_execute("UPDATE users SET subscription=?, updated_at=? WHERE id=?", (req.tier, _utcnow(), req.user_id))
    await _push_notification(req.user_id, "Plan Updated", f"Your plan was updated to {req.tier.upper()} by an admin.", "success", actor_id=admin["sub"])
    return {"success": True}

@app.post("/admin/users/{uid}/impersonate")
async def admin_impersonate(uid: str, admin: Dict = Depends(require_admin)) -> Dict:
    target = await db_fetchone("SELECT id,email,role,subscription FROM users WHERE id=?", (uid,))
    if not target: raise HTTPException(404, "User not found")
    token = _create_jwt({"sub": uid, "email": target["email"], "role": target["role"], "subscription": target["subscription"], "impersonated_by": admin["sub"]}, 60)
    return {"access_token": token, "warning": "1-hour impersonation token"}

@app.get("/admin/stats")
async def admin_stats(admin: Dict = Depends(require_admin)) -> Dict:
    return {
        "users":      await db_count("SELECT COUNT(*) FROM users"),
        "messages":   await db_count("SELECT COUNT(*) FROM chat_history"),
        "documents":  await db_count("SELECT COUNT(*) FROM documents"),
        "agent_jobs": await db_count("SELECT COUNT(*) FROM agent_jobs"),
        "connectors": await db_count("SELECT COUNT(*) FROM connectors"),
        "websites":   await db_count("SELECT COUNT(*) FROM websites"),
        "by_tier": {
            "free":       await db_count("SELECT COUNT(*) FROM users WHERE subscription='free'"),
            "pro":        await db_count("SELECT COUNT(*) FROM users WHERE subscription='pro'"),
            "enterprise": await db_count("SELECT COUNT(*) FROM users WHERE subscription='enterprise'"),
        },
    }

@app.post("/admin/notify")
async def admin_notify(req: NotifCreate, admin: Dict = Depends(require_admin)) -> Dict:
    await _push_notification(req.user_id, req.title, req.message, req.type, req.link, req.is_broadcast, admin["sub"])
    return {"success": True}

@app.post("/admin/coupons")
async def admin_create_coupon(req: CouponCreate, admin: Dict = Depends(require_admin)) -> Dict:
    exp = None
    if req.expires_days: exp = (datetime.now(timezone.utc) + timedelta(days=req.expires_days)).isoformat()
    cid = _new_id()
    await db_execute(
        "INSERT INTO coupons (id,code,grants_tier,duration_days,max_uses,created_by,expires_at,created_at)"
        " VALUES (?,?,?,?,?,?,?,?)",
        (cid, req.code, req.grants_tier, req.duration_days, req.max_uses, admin["sub"], exp, _utcnow()),
    )
    return {"id": cid, "code": req.code}

@app.delete("/admin/users/{uid}")
async def admin_delete_user(uid: str, admin: Dict = Depends(require_admin)) -> Dict:
    if uid == admin["sub"]: raise HTTPException(400, "Cannot delete your own account")
    await db_execute("DELETE FROM users WHERE id=?", (uid,))
    return {"success": True}

@app.get("/logs")
async def get_logs(admin: Dict = Depends(require_admin)) -> Dict:
    try: return {"logs": Path("jazz.log").read_text(encoding="utf-8").splitlines()[-500:]}
    except FileNotFoundError: return {"logs": []}


# ══════════════════════════════════════════════════════════════════════════════
# §42  GDPR
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/gdpr/export")
async def gdpr_export(user: Dict = Depends(get_current_user)) -> StreamingResponse:
    uid = user["sub"]
    export = {
        "export_generated_at": _utcnow(),
        "profile":    await db_fetchone("SELECT id,email,full_name,role,subscription,timezone,created_at FROM users WHERE id=?", (uid,)),
        "sessions":   await db_fetchall("SELECT id,title,turn_count,created_at FROM chat_sessions WHERE user_id=?", (uid,)),
        "messages":   await db_fetchall("SELECT session_id,role,content,created_at FROM chat_history WHERE user_id=? AND is_hidden=0 ORDER BY created_at", (uid,)),
        "memories":   await db_fetchall("SELECT key,value,source,confidence FROM memories WHERE user_id=? AND is_active=1", (uid,)),
        "documents":  await db_fetchall("SELECT original_name,file_type,file_size_bytes FROM documents WHERE user_id=?", (uid,)),
    }
    content  = json.dumps(export, indent=2, default=str)
    filename = f"jazz_export_{uid[:8]}_{datetime.now().strftime('%Y%m%d')}.json"
    return StreamingResponse(
        io.StringIO(content), media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# §43  HEALTH & INFO
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/health")
async def health() -> Dict:
    db_ok = False
    try: await db_fetchone("SELECT 1"); db_ok = True
    except Exception: pass
    uptime = int((datetime.now(timezone.utc) - _SERVER_START).total_seconds())
    return {
        "status":   "healthy" if db_ok else "degraded",
        "version":  "11.0.0",
        "uptime_s": uptime,
        "services": {
            "sqlite":    db_ok,
            "chromadb":  _chroma_client is not None,
            "scheduler": _scheduler is not None and _scheduler.running,
            "livekit":   _lk_available(),
        },
        "features": [
            "auto_thinking_mode", "full_system_access", "livekit_voice",
            "multilingual_stt", "streaming_sse", "hybrid_rag", "reranking",
            "rolling_summaries", "tiered_memory", "agent_loop", "cron_jobs",
            "branching", "vision", "connectors", "website_builder", "gdpr",
        ],
    }

@app.get("/model-info")
async def model_info() -> Dict:
    return {
        "models":          [{"id": k, "label": v["label"]} for k, v in _MODEL_REGISTRY.items()],
        "styles":          list(_STYLE_HINTS.keys()),
        "tools":           list(_TOOL_REGISTRY.keys()),
        "connector_types": ["slack", "google_sheets", "excel", "power_bi", "http"],
        "livekit_url":     LIVEKIT_URL,
        "livekit_ready":   _lk_available(),
    }

@app.get("/")
async def root() -> HTMLResponse:
    frontend = Path("frontend/index.html")
    if frontend.exists():
        return HTMLResponse(frontend.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>JAZZ AI v11 — API running ✅</h1><p>See <a href='/docs'>/docs</a>.</p>")


# ══════════════════════════════════════════════════════════════════════════════
# §44  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "server11:app",
        host="0.0.0.0",
        port=int(os.getenv("PORT", "8000")),
        reload=False,
        workers=1,
        log_level="info",
        access_log=True,
        timeout_keep_alive=75,
    )